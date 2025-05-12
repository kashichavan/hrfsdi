from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.core.cache import cache
from django.conf import settings
from django.utils import timezone
from .models import Student, Requirement, RequirementStudent
from .forms import RequirementForm, RequirementEditForm
import pandas as pd
import re
import threading
import uuid
import json
import os
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

from student_data import models

# Utility functions
def normalize_column_name(name):
    name = str(name).strip().lower()
    name = re.sub(r'\s+', '_', name)
    name = name.replace('%', '_percent')
    name = name.replace('10th', 'tenth')
    name = name.replace('12th', 'twelfth')
    return name

def process_excel_data(file_path, task_id):
    try:
        cache.set(f'task_{task_id}_status', json.dumps({
            'status': 'processing',
            'progress': 0,
            'message': 'Starting processing...'
        }), timeout=3600)

        df = pd.read_excel(file_path)

        if df.empty:
            raise ValueError("Uploaded file is empty.")

        df.columns = [normalize_column_name(col) for col in df.columns]
        
        # Handle case where 'sl_no' might not exist
        if 'sl_no' in df.columns:
            df = df.drop(columns=['sl_no'])

        required_columns = {
            'name', 'contact_number', 'degree', 'stream',
            'yop', 'tenth_percent', 'twelfth_percent',
            'degree_percent', 'gender', 'type_of_data'
        }
        missing = required_columns - set(df.columns)
        if missing:
            raise ValueError(f"Missing columns: {', '.join(missing)}")

        df['yop'] = pd.to_numeric(df['yop'], errors='coerce').fillna(0).astype(int)
        numeric_cols = ['tenth_percent', 'twelfth_percent', 'degree_percent']
        df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors='coerce')

        df = df.drop_duplicates(subset=['contact_number'], keep='first')
        
        total_rows = len(df)
        chunk_size = 1000
        processed_rows = 0
        duplicates_count = 0

        with transaction.atomic():
            for start in range(0, total_rows, chunk_size):
                end = min(start + chunk_size, total_rows)
                chunk_df = df.iloc[start:end]
                
                for _, row in chunk_df.iterrows():
                    if pd.isna(row.get('name', '')) or pd.isna(row.get('contact_number', '')):
                        continue
                        
                    contact_number = str(row.get('contact_number', ''))
                    
                    try:
                        existing_student = Student.objects.get(contact_number=contact_number)
                        existing_student.name = row.get('name', '')
                        existing_student.degree = row.get('degree', '')
                        existing_student.stream = row.get('stream', '')
                        existing_student.yop = row.get('yop', 0)
                        existing_student.tenth_percent = row.get('tenth_percent', 0.0)
                        existing_student.twelfth_percent = row.get('twelfth_percent', 0.0)
                        existing_student.degree_percent = row.get('degree_percent', 0.0)
                        existing_student.gender = row.get('gender', 'Male')
                        existing_student.type_of_data = row.get('type_of_data', '')
                        existing_student.save()
                        duplicates_count += 1
                    except Student.DoesNotExist:
                        Student.objects.create(
                            name=row.get('name', ''),
                            contact_number=contact_number,
                            degree=row.get('degree', ''),
                            stream=row.get('stream', ''),
                            yop=row.get('yop', 0),
                            tenth_percent=row.get('tenth_percent', 0.0),
                            twelfth_percent=row.get('twelfth_percent', 0.0),
                            degree_percent=row.get('degree_percent', 0.0),
                            gender=row.get('gender', 'Male'),
                            type_of_data=row.get('type_of_data', '')
                        )

                processed_rows += len(chunk_df)
                progress = min(int((processed_rows / total_rows) * 100), 95)
                cache.set(f'task_{task_id}_status', json.dumps({
                    'status': 'processing',
                    'progress': progress,
                    'message': f'Processed {processed_rows} / {total_rows} rows. Updated {duplicates_count} existing records.'
                }), timeout=3600)

        cache.set(f'task_{task_id}_status', json.dumps({
            'status': 'completed',
            'progress': 100,
            'message': f'Successfully processed {processed_rows} records. Updated {duplicates_count} existing records.'
        }), timeout=3600)

    except Exception as e:
        cache.set(f'task_{task_id}_status', json.dumps({
            'status': 'error',
            'message': str(e)
        }), timeout=3600)
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)

# Student views
@login_required
def upload_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if not excel_file or not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Invalid file format. Please upload a valid Excel file (.xlsx or .xls)')
            return render(request, 'upload.html', {'error': 'Invalid file format'})

        try:
            temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_excel')
            os.makedirs(temp_dir, exist_ok=True)

            unique_filename = f"{uuid.uuid4().hex}_{excel_file.name}"
            file_path = os.path.join(temp_dir, unique_filename)

            with open(file_path, 'wb+') as destination:
                for chunk in excel_file.chunks():
                    destination.write(chunk)

            task_id = str(uuid.uuid4())

            cache.set(f'task_{task_id}_status', json.dumps({
                'status': 'pending',
                'progress': 0,
                'message': 'Task pending...'
            }), timeout=3600)

            threading.Thread(
                target=process_excel_data,
                args=(file_path, task_id),
                daemon=True
            ).start()

            request.session['excel_task_id'] = task_id
            return redirect('student_data:processing_page')

        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
            return render(request, 'upload.html', {'error': f"Error processing file: {str(e)}"})

    return render(request, 'upload.html')

@login_required
def processing_page(request):
    task_id = request.session.get('excel_task_id')
    if not task_id:
        messages.error(request, "No active upload task found")
        return redirect('student_data:upload_excel')
    return render(request, 'processing.html', {'task_id': task_id})

@login_required
def check_task_status(request):
    task_id = request.GET.get('task_id')
    if not task_id:
        return JsonResponse({'status': 'error', 'message': 'No task ID provided'})

    status_data = cache.get(f'task_{task_id}_status')
    if not status_data:
        return JsonResponse({'status': 'error', 'message': 'Task not found or expired'})

    try:
        status = json.loads(status_data)
        return JsonResponse(status)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid task status data'})


@login_required
def download_excel_template(request):
    # Create a template Excel file with required columns
    columns = [
        'sl_no', 'name', 'contact_number', 'degree', 'stream',
        'yop', 'tenth_percent', 'twelfth_percent',
        'degree_percent', 'gender', 'type_of_data'
    ]
    
    # Create a sample data row
    sample_data = [{
        'sl_no': 1,
        'name': 'John Doe',
        'contact_number': '9876543210',
        'degree': 'B.Tech',
        'stream': 'Computer Science',
        'yop': 2023,
        'tenth_percent': 85.5,
        'twelfth_percent': 78.9,
        'degree_percent': 82.3,
        'gender': 'Male',
        'type_of_data': 'Student'
    }]
    
    df = pd.DataFrame(sample_data, columns=columns)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=student_upload_template.xlsx'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='StudentDataTemplate')

    return response


from django.shortcuts import render
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Q, Count
from django.contrib.auth.decorators import login_required
from .models import Student, RequirementStudent

@login_required
def student_list(request):
    # Get filter parameters from request
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort', '-scheduled_requirements')  # Default to high-to-low
    page_number = request.GET.get('page', 1)
    type_filter = request.GET.get('type_filter', '')
    degree_filter = request.GET.get('degree', '')
    stream_filters = request.GET.getlist('stream')
    min_tenth = request.GET.get('min_tenth', '')
    min_twelfth = request.GET.get('min_twelfth', '')
    min_degree = request.GET.get('min_degree', '')
    gender_filter = request.GET.get('gender', '')
    min_yop = request.GET.get('min_yop', '')
    max_yop = request.GET.get('max_yop', '')

    # Start with base queryset
    students = Student.objects.all()
    students = students.exclude(requirementstudent__status='selected')
    students = students.exclude(is_placed=True)

    # Apply type_of_data filter
    if type_filter:
        type_filter_lower = type_filter.lower()
        if type_filter_lower != 'all':
            if type_filter_lower == 'placement_activity':
                students = students.filter(
                    Q(type_of_data__iexact='placement_activity') | 
                    Q(type_of_data__iexact='placement activity')
                )
            elif type_filter_lower == 'early_placement':
                students = students.filter(
                    Q(type_of_data__iexact='early_placement') | 
                    Q(type_of_data__iexact='early placement')
                )
            elif type_filter_lower == 'dropout':
                students = students.filter(is_dropout=True)
            else:
                students = students.filter(type_of_data__iexact=type_filter_lower)

    # Filter by degree (case-insensitive)
    if degree_filter:
        students = students.filter(degree__iexact=degree_filter)

    # Filter by stream(s) (case-insensitive)
    if stream_filters:
        stream_query = Q()
        for stream in stream_filters:
            if stream:
                stream_query |= Q(stream__iexact=stream)
        if stream_query:
            students = students.filter(stream_query)

    # Filter by percentages and year
    try:
        if min_tenth:
            students = students.filter(tenth_percent__gte=float(min_tenth))
        if min_twelfth:
            students = students.filter(twelfth_percent__gte=float(min_twelfth))
        if min_degree:
            students = students.filter(degree_percent__gte=float(min_degree))
        if min_yop:
            students = students.filter(yop__gte=int(min_yop))
        if max_yop:
            students = students.filter(yop__lte=int(max_yop))
    except ValueError:
        pass  # Ignore bad inputs

    # Filter by gender
    if gender_filter:
        students = students.filter(gender__iexact=gender_filter)

    # Search filter
    if search_query:
        students = students.filter(
            Q(name__icontains=search_query) |
            Q(contact_number__icontains=search_query)
        )

    # Sorting configuration
    valid_sorts = {
        'name', '-name',
        'yop', '-yop',
        'tenth_percent', '-tenth_percent',
        'twelfth_percent', '-twelfth_percent',
        'degree_percent', '-degree_percent',
        'total_requirements', '-total_requirements',
        'scheduled_requirements', '-scheduled_requirements',
        'gender', '-gender',
        'type_of_data', '-type_of_data',
        'created_at', '-created_at'
    }

    if sort_by in valid_sorts:
        if sort_by not in ['scheduled_requirements', '-scheduled_requirements']:
            students = students.order_by('-scheduled_requirements', sort_by, 'name')
        else:
            students = students.order_by(sort_by, 'name')
    else:
        students = students.order_by('-scheduled_requirements', 'name')

    # Get cleaned unique values for filters
    def get_cleaned_unique_values(field_name):
        values = Student.objects.exclude(**{f'{field_name}__isnull': True}) \
                               .exclude(**{f'{field_name}__exact': ''}) \
                               .values_list(field_name, flat=True) \
                               .distinct()
        
        cleaned = set()
        for value in values:
            if value:
                clean_val = ' '.join(word.strip().title() for word in str(value).split())
                cleaned.add(clean_val)
        return sorted(cleaned)

    unique_degrees = get_cleaned_unique_values('degree')
    unique_streams = get_cleaned_unique_values('stream')

    # Pagination
    paginator = Paginator(students, 50)
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.get_page(1)
    except EmptyPage:
        page_obj = paginator.get_page(paginator.num_pages)

    # Standardize display values
    for student in page_obj:
        if student.name:
            student.display_name = ' '.join(word.title() for word in student.name.split())
        if student.degree:
            student.display_degree = ' '.join(word.title() for word in student.degree.split())
        if student.stream:
            student.display_stream = ' '.join(word.title() for word in student.stream.split())

    # Type counts for quick filters
    base_query = Student.objects.exclude(requirementstudent__status='selected').exclude(is_placed=True)
    type_counts = {
        'all': base_query.count(),
        'fsdi': base_query.filter(type_of_data__iexact='fsdi').count(),
        'super100': base_query.filter(type_of_data__iexact='super100').count(),
        'tuition': base_query.filter(type_of_data__iexact='tuition').count(),
        'legend': base_query.filter(type_of_data__iexact='legend').count(),
        'placement_activity': base_query.filter(
            Q(type_of_data__iexact='placement_activity') | 
            Q(type_of_data__iexact='placement activity')
        ).count(),
        'early_placement': base_query.filter(
            Q(type_of_data__iexact='early_placement') | 
            Q(type_of_data__iexact='early placement')
        ).count(),
        'dropout': base_query.filter(is_dropout=True).count(),
    }

    # Context data
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'sort_by': sort_by,
        'type_filter': type_filter,
        'total_count': students.count(),
        'type_counts': type_counts,
        'degree_filter': degree_filter,
        'stream_filters': stream_filters,
        'min_tenth': min_tenth,
        'min_twelfth': min_twelfth,
        'min_degree': min_degree,
        'gender_filter': gender_filter,
        'min_yop': min_yop,
        'max_yop': max_yop,
        'unique_degrees': unique_degrees,
        'unique_streams': unique_streams,
        'current_sort': sort_by,
        'sort_options': [
            ('-scheduled_requirements', 'Scheduled (High to Low)'),
            ('scheduled_requirements', 'Scheduled (Low to High)'),
            ('-total_requirements', 'Total (High to Low)'),
            ('total_requirements', 'Total (Low to High)'),
            ('name', 'Name (A-Z)'),
            ('-name', 'Name (Z-A)'),
            ('yop', 'YOP (Oldest)'),
            ('-yop', 'YOP (Newest)'),
            ('tenth_percent', '10th % (Low)'),
            ('-tenth_percent', '10th % (High)'),
            ('twelfth_percent', '12th % (Low)'),
            ('-twelfth_percent', '12th % (High)'),
            ('degree_percent', 'Degree % (Low)'),
            ('-degree_percent', 'Degree % (High)'),
        ]
    }

    return render(request, 'student_list.html', context)

@login_required
def student_detail(request, student_id):
    student = get_object_or_404(
        Student.objects.annotate(
            total_requirements_count=Count('requirementstudent'),
            scheduled_requirements_count=Count(
                'requirementstudent',
                filter=Q(requirementstudent__requirement__is_scheduled=True)
            )
        ),
        pk=student_id
    )

    student_requirements = (
        RequirementStudent.objects
        .filter(student=student)
        .select_related('requirement')
        .order_by('-requirement__created_at')
    )

    # Get subject ratings for the student
    subject_ratings = (
        StudentSubjectRating.objects
        .filter(student=student)
        .select_related('subject')
        .order_by('subject__name')
    )

    context = {
        'student': student,
        'student_requirements': student_requirements,
        'subject_ratings': subject_ratings,
        'total_requirements': student.total_requirements_count,
        'scheduled_requirements': student.scheduled_requirements_count,
        'pending_requirements_count': student.total_requirements_count - student.scheduled_requirements_count,
        'js_data': {
            'name': student.name,
            'degree': student.degree,
            'stream': student.stream,
            'degree_percent': float(getattr(student, 'degree_percent', 0.0)),
            'tenth_percent': float(getattr(student, 'tenth_percent', 0.0)),
            'twelfth_percent': float(getattr(student, 'twelfth_percent', 0.0)),
            'total_requirements': student.total_requirements_count,
            'scheduled_requirements': student.scheduled_requirements_count,
        }
    }

    return render(request, 'student_detail.html', context)
@login_required
def requirement_list(request):
    # Get search and filter parameters
    company_name = request.GET.get('company_name', '')
    month = request.GET.get('month', '')
    sort_by = request.GET.get('sort', '-requirement_date')
    
    # Start with base queryset
    requirements = Requirement.objects.all()
    
    # Apply filters
    if company_name:
        requirements = requirements.filter(company_name__icontains=company_name)
    
    if month:
        year, month_num = map(int, month.split('-'))
        requirements = requirements.filter(
            requirement_date__year=year,
            requirement_date__month=month_num
        )
    
    # Annotate with student count and order
    requirements = requirements.order_by(sort_by).annotate(
        student_count=Count('students')
    )
    
    # Get counts for stats
    total_count = requirements.count()
    
    # Pagination
    paginator = Paginator(requirements, 10)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    context = {
        'page_obj': page_obj,
        'company_name': company_name,
        'month': month,
        'sort_by': sort_by,
        'total_count': total_count,
    }
    
    return render(request, 'requirement_list.html', context)


from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from .forms import RequirementForm
from .models import Student, RequirementStudent, RequirementSubject, Subject
import pandas as pd
@login_required
@transaction.atomic
def add_requirement(request):
    if request.method == 'POST':
        form = RequirementForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Save main requirement object
                requirement = form.save(commit=False)
                requirement.created_by = request.user
                requirement.schedule_status = 'scheduled' if requirement.is_scheduled else 'not_scheduled'
                requirement.save()
                form.save_m2m()  # Save many-to-many relationships

                # Handle subjects and other subject name
                subjects = form.cleaned_data['subjects']
                other_subject_name = form.cleaned_data.get('other_subject_name', '').strip()
                other_subject = Subject.objects.filter(name__iexact='other').first()

                # Process subjects and handle "Other" replacement
                final_subjects = []
                
                if other_subject and other_subject in subjects:
                    # Validate other subject name
                    if not other_subject_name:
                        form.add_error('other_subject_name', 'Please specify the subject name when selecting "Other"')
                        raise ValidationError("Missing other subject name")
                    
                    # Save other subject name separately
                    requirement.other_subject_name = other_subject_name
                else:
                    final_subjects = list(subjects)
                
                # Create RequirementSubject entries
                for subject in final_subjects:
                    RequirementSubject.objects.get_or_create(
                        requirement=requirement,
                        subject=subject
                    )

                # Process student Excel file
                if 'student_file' in request.FILES:
                    return handle_excel_upload(request, requirement, form)
                
                messages.success(request, 'Requirement created successfully!')
                return redirect('student_data:requirement_detail', requirement.id)

            except Exception as e:
                messages.error(request, f'Error creating requirement: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = RequirementForm()

    return render(request, 'add_requirement.html', {
        'form': form,
        'title': 'Add New Requirement'
    })
def handle_excel_upload(request, requirement, form):
    """Helper function to process student Excel file"""
    try:
        excel_file = request.FILES['student_file']
        df = pd.read_excel(excel_file)
        mobile_column = form.cleaned_data.get('mobile_column', 'mobile').strip()
        
        if mobile_column not in df.columns:
            messages.error(request, f"Column '{mobile_column}' not found in the Excel file.")
            return redirect('student_data:add_requirement')

        processing_results = {
            'total': 0,
            'added': 0,
            'existing': 0,
            'invalid': 0
        }

        for index, row in df.iterrows():
            processing_results['total'] += 1
            mobile = str(row[mobile_column]).strip()
            
            # Validate mobile number
            if not mobile or mobile.lower() == 'nan':
                processing_results['invalid'] += 1
                continue
            
            # Clean mobile number
            clean_mobile = ''.join(filter(str.isdigit, mobile))
            if len(clean_mobile) < 10:
                processing_results['invalid'] += 1
                continue
            clean_mobile = clean_mobile[-10:]  # Take last 10 digits

            # Find matching students
            students = Student.objects.filter(contact_number__endswith=clean_mobile)
            if not students.exists():
                processing_results['invalid'] += 1
                continue
                
            # Add valid students
            for student in students:
                _, created = RequirementStudent.objects.get_or_create(
                    requirement=requirement,
                    student=student
                )
                if created:
                    processing_results['added'] += 1
                else:
                    processing_results['existing'] += 1

        # Build result message
        msg_parts = [
            f"Processed {processing_results['total']} students:",
            f"Added {processing_results['added']} new entries",
            f"Skipped {processing_results['existing']} duplicates",
            f"{processing_results['invalid']} invalid entries"
        ]
        messages.success(request, ' '.join(msg_parts))
        
        return redirect('student_data:requirement_detail', requirement.id)

    except Exception as e:
        requirement.delete()  # Rollback if error occurs
        messages.error(request, f'Error processing Excel file: {str(e)}')
        return redirect('student_data:add_requirement')
from django.utils.http import url_has_allowed_host_and_scheme
from django.shortcuts import get_object_or_404, render
from .models import Requirement, RequirementStudent, RequirementSubject

def requirement_detail(request, pk):
    requirement = get_object_or_404(
        Requirement.objects.select_related('scheduled_details')
        .prefetch_related('subjects', 'students'),
        pk=pk
    )
    
    # Get all requirement-student relationships
    requirement_students = RequirementStudent.objects.filter(
        requirement=requirement
    ).select_related('student')
    
    # Get academic percentages from the requirement
    academic_percentages = {
        '10th': requirement.percentage_10th,
        '12th': requirement.percentage_12th,
        'degree': requirement.percentage_degree,
        'masters': requirement.percentage_master
    }
    
    # Get subjects with their optional custom names
    requirement_subjects = RequirementSubject.objects.filter(
        requirement=requirement
    ).select_related('subject')
    
    # Student statistics
    total_students = requirement_students.count()
    status_counts = {
        'selected': requirement_students.filter(status='selected').count(),
        'rejected': requirement_students.filter(status='rejected').count(),
        'pending': requirement_students.filter(status='pending').count(),
        'on_hold': requirement_students.filter(status='on_hold').count(),
    }

    context = {
        'requirement': requirement,
        'requirement_students': requirement_students,
        'academic_percentages': academic_percentages,
        'requirement_subjects': requirement_subjects,
        'total_students': total_students,
        'selected_students': status_counts['selected'],
        'rejected_students': status_counts['rejected'],
        'pending_students': status_counts['pending'],
        'on_hold_students': status_counts['on_hold'],
        'previous_url': request.META.get('HTTP_REFERER'),
    }
    
    return render(request, 'requirement_detail.html', context)

import io
import pandas as pd
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from .models import Requirement, RequirementStudent


def export_requirement_students_excel(request, pk):
    requirement = get_object_or_404(Requirement, pk=pk)
    requirement_students = RequirementStudent.objects.filter(requirement=requirement).select_related('student')

    if not requirement_students.exists():
        return HttpResponse("No students assigned to this requirement.", status=404)

    # Build student data
    student_data = []
    for req_student in requirement_students:
        student = req_student.student
        if not student:
            continue
        student_data.append({
            'Student Name': student.name,
            'Contact Number': student.contact_number,
            'Gender': student.gender,
            'Degree': student.degree,
            'Stream': student.stream,
            'Year of Passing': student.yop,
            '10th %': student.tenth_percent,
            '12th %': student.twelfth_percent,
            'Degree %': student.degree_percent,
            'Type of Data': student.type_of_data,
            'Status': req_student.get_status_display(),
            'Feedback': req_student.feedback or '',
            'Is Placed': "Yes" if student.is_placed else "No"
        })

    df = pd.DataFrame(student_data)

    # Excel file creation
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Assigned Students')

        worksheet = writer.sheets['Assigned Students']
        workbook = writer.book

        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="DDDDDD")

        for col_idx, column in enumerate(df.columns, 1):
            max_length = max(df[column].astype(str).map(len).max(), len(column))
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = min(max_length + 4, 40)
            cell = worksheet[f"{col_letter}1"]
            cell.font = header_font
            cell.fill = header_fill

    excel_buffer.seek(0)
    filename = f"{requirement.company_name.replace(' ', '_')}_students_{requirement.id}.xlsx"
    response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
# views.py
def update_student_feedback(request, requirement_id):
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        feedback = request.POST.get('feedback', '')
        
        try:
            rs = RequirementStudent.objects.get(
                requirement_id=requirement_id,
                student_id=student_id
            )
            rs.feedback = feedback
            rs.save()
            messages.success(request, 'Feedback updated successfully')
        except RequirementStudent.DoesNotExist:
            messages.error(request, 'Student not found in this requirement')
        
        return redirect('student_data:requirement_detail',pk=requirement_id)

from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from .models import Requirement
from .forms import RequirementEditForm

@login_required
@transaction.atomic
def requirement_edit(request, pk):
    requirement = get_object_or_404(Requirement, pk=pk)

    if request.method == 'POST':
        form = RequirementEditForm(request.POST, instance=requirement)
        if form.is_valid():
            try:
                requirement = form.save(commit=False)
                requirement.schedule_status = 'scheduled' if requirement.is_scheduled else 'not_scheduled'
                requirement.modified_by = request.user
                requirement.save()

                students = list(requirement.students.all())
                for student in students:
                    student.update_requirement_counts()

                Student.objects.bulk_update(
                    students,
                    ['scheduled_count', 'not_scheduled_count', 'modified_at'],
                    batch_size=100
                )

                messages.success(request, 'Requirement updated successfully!')
                return redirect('student_data:requirement_detail', pk=requirement.id)

            except Exception as e:
                messages.error(request, f'Error updating requirement: {str(e)}')
                # Optional: log error using logging module
                # import logging; logging.exception(e)
    else:
        form = RequirementEditForm(instance=requirement)

    return render(request, 'requirement_edit.html', {
        'form': form,
        'requirement': requirement,
        'title': 'Edit Requirement'
    })
@login_required
def requirement_students(request, pk):
    requirement = get_object_or_404(Requirement, pk=pk)
    assigned_requirement_students = RequirementStudent.objects.filter(requirement=requirement)
    assigned_student_ids = list(assigned_requirement_students.values_list('student_id', flat=True))

    available_students = Student.objects.all()

    filters = {
        'search': request.GET.get('search', ''),
        'degrees': request.GET.getlist('degree', []),
        'streams': request.GET.getlist('stream', []),
        'gender': request.GET.get('gender', ''),
        'min_percentage': request.GET.get('min_percentage', ''),
        'min_yop': request.GET.get('min_yop', ''),
        'max_yop': request.GET.get('max_yop', ''),
        'show_assigned': request.GET.get('show_assigned', 'false') == 'true'
    }

    query = Q()

    if filters['search']:
        query |= (Q(name__icontains=filters['search']) |
                 Q(contact_number__icontains=filters['search']) |
                 Q(email__icontains=filters['search']))

    if filters['degrees']:
        query &= Q(degree__in=filters['degrees'])

    if filters['streams']:
        query &= Q(stream__in=filters['streams'])

    if filters['gender']:
        query &= Q(gender=filters['gender'])

    if filters['min_percentage']:
        try:
            query &= Q(degree_percent__gte=float(filters['min_percentage']))
        except ValueError:
            pass

    if filters['min_yop']:
        try:
            query &= Q(yop__gte=int(filters['min_yop']))
        except ValueError:
            pass

    if filters['max_yop']:
        try:
            query &= Q(yop__lte=int(filters['max_yop']))
        except ValueError:
            pass

    available_students = available_students.filter(query)

    if not filters['show_assigned']:
        available_students = available_students.exclude(id__in=assigned_student_ids)

    if request.method == 'POST':
        action = request.POST.get('action')
        student_ids = request.POST.getlist('student_ids')

        if action == 'add' and student_ids:
            added_count = 0
            existing_count = 0
            for sid in student_ids:
                try:
                    student = Student.objects.get(pk=sid)
                    _, created = RequirementStudent.objects.get_or_create(
                        requirement=requirement,
                        student=student,
                        defaults={'status': 'pending'}
                    )
                    if created:
                        added_count += 1
                    else:
                        existing_count += 1
                except Student.DoesNotExist:
                    continue

            if added_count > 0:
                messages.success(request, f'{added_count} students successfully added.')
            if existing_count > 0:
                messages.info(request, f'{existing_count} students were already assigned.')
            
            return redirect('student_data:requirement_students', pk=requirement.id)

        elif action == 'update_status' and student_ids:
            new_status = request.POST.get('new_status')
            if new_status in ['pending', 'selected', 'rejected', 'on_hold']:
                updated = RequirementStudent.objects.filter(
                    requirement=requirement,
                    student_id__in=student_ids
                ).update(status=new_status)
                messages.success(request, f'Status updated for {updated} students.')
            
            return redirect('requirement_students', pk=requirement.id)

    page_size = int(request.GET.get('page_size', 12))
    paginator = Paginator(available_students.order_by('-created_at'), page_size)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    unique_degrees = Student.objects.values_list('degree', flat=True).distinct().order_by('degree')
    unique_streams = Student.objects.values_list('stream', flat=True).distinct().order_by('stream')

    context = {
        'requirement': requirement,
        'page_obj': page_obj,
        'assigned_students': assigned_requirement_students.select_related('student'),
        'filters': filters,
        'unique_degrees': unique_degrees,
        'unique_streams': unique_streams,
        'total_available': available_students.count(),
        'total_assigned': assigned_requirement_students.count(),
        'page_size_options': [12, 24, 48, 96],
        'current_page_size': page_size,
    }

    return render(request, 'requirement_students.html', context)
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from django.db import transaction
from .models import Requirement, RequirementStudent, ScheduledRequirement

def update_requirement_schedule(request, requirement_id):
    requirement = get_object_or_404(Requirement, id=requirement_id)
    
    if request.method == 'POST':
        status = request.POST.get('status')
        schedule_date = request.POST.get('schedule_date')
        
        try:
            with transaction.atomic():
                old_status = requirement.schedule_status
                
                if status == 'scheduled' and schedule_date:
                    # Handle scheduling case
                    requirement.is_scheduled = True
                    requirement.schedule_date = schedule_date
                    requirement.schedule_status = 'scheduled'
                    requirement.save()
                    
                    # Get or create scheduled details
                    scheduled_details, created = ScheduledRequirement.objects.get_or_create(
                        requirement=requirement,
                        defaults={'scheduled_date': schedule_date}
                    )
                    if not created and scheduled_details.scheduled_date != schedule_date:
                        scheduled_details.scheduled_date = schedule_date
                        scheduled_details.save()
                    
                    # Update student statuses
                    requirement.requirementstudent_set.filter(status='pending').update(status='on_hold')
                    
                else:
                    # Handle unscheduling case
                    requirement.is_scheduled = False
                    requirement.schedule_date = None
                    requirement.schedule_status = 'not_scheduled'
                    requirement.save()
                    
                    # More robust deletion of scheduled details
                    try:
                        if hasattr(requirement, 'scheduled_details'):
                            # First check if the object exists in database
                            if ScheduledRequirement.objects.filter(requirement=requirement).exists():
                                requirement.scheduled_details.delete()
                    except Exception as e:
                        # If deletion fails, just continue - it's not critical
                        messages.warning(request, f'Could not clean up schedule details: {str(e)}')
                    
                    # Reset student statuses
                    requirement.requirementstudent_set.filter(status='on_hold').update(status='pending')
                
                messages.success(request, 'Requirement schedule updated successfully')
                return redirect('student_data:requirement_detail', pk=requirement.id)
                
        except Exception as e:
            messages.error(request, f'Error updating schedule: {str(e)}')
            return redirect('student_data:requirement_detail', pk=requirement.id)
    
    return redirect('student_data:requirement_detail', pk=requirement.id)
from django.shortcuts import get_object_or_404, redirect
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import Requirement, ScheduledRequirement, RequirementStudent, Student

@require_POST
def update_drive_result(request, pk):
    requirement = get_object_or_404(Requirement, id=pk)
    feedback = request.POST.get('feedback', '')
    selected_student_ids = request.POST.getlist('selected_students', [])

    scheduled_detail, _ = ScheduledRequirement.objects.get_or_create(requirement=requirement)
    scheduled_detail.feedback = feedback
    scheduled_detail.save()

    selected_students = Student.objects.filter(id__in=selected_student_ids)
    all_students = requirement.students.all()

    for student in all_students:
        status = 'selected' if student in selected_students else 'rejected'
        RequirementStudent.objects.update_or_create(
            requirement=requirement,
            student=student,
            defaults={'status': status}
        )

    update_drive_result_status(requirement)
    messages.success(request, "Drive result updated successfully.")
    return redirect('student_data:requirement_detail', pk=pk)

@login_required
def update_student_status(request, requirement_id, student_id):
    requirement = get_object_or_404(Requirement, pk=requirement_id)
    student = get_object_or_404(Student, pk=student_id)
    req_student = get_object_or_404(RequirementStudent, requirement=requirement, student=student)

    if request.method == 'POST':
        new_status = request.POST.get('status', '').strip()
        valid_statuses = ['pending', 'selected', 'rejected']
        old_status = req_student.status

        if new_status in valid_statuses:
            req_student.status = new_status
            req_student.save()

            # Update placement flag on Student
            if old_status == 'selected' and new_status != 'selected':
                still_selected_elsewhere = RequirementStudent.objects.filter(
                    student=student,
                    status='selected'
                ).exclude(requirement=requirement).exists()
                if not still_selected_elsewhere:
                    student.is_placed = False
                    student.placed_date = None
                    student.placed_company = None
                    student.save()
            elif new_status == 'selected':
                student.is_placed = True
                student.placed_date = timezone.now().date()
                student.placed_company = requirement.company_name
                student.save()

            update_drive_result_status(requirement)
            messages.success(request, f"Status updated for {student.name}.")
        else:
            messages.error(request, "Invalid status.")
    
    return redirect('student_data:requirement_detail', pk=requirement_id)

def update_drive_result_status(requirement):
    """Update ScheduledRequirement result and students_appeared based on RequirementStudent statuses."""
    if not hasattr(requirement, 'scheduled_detail'):
        scheduled_detail, _ = ScheduledRequirement.objects.get_or_create(requirement=requirement)
    else:
        scheduled_detail = requirement.scheduled_detail

    selected_students = Student.objects.filter(
        requirementstudent__requirement=requirement,
        requirementstudent__status='selected'
    )
    scheduled_detail.students_appeared.set(selected_students)

    # Recalculate result
    all_statuses = RequirementStudent.objects.filter(requirement=requirement).values_list('status', flat=True)

    if all(status == 'pending' for status in all_statuses):
        result = 'pending'
    elif any(status == 'selected' for status in all_statuses):
        result = 'selected'
    else:
        result = 'no_selects'

    if scheduled_detail.result != result:
        scheduled_detail.result = result
        scheduled_detail.save()




@login_required
def bulk_import_students(request, requirement_id):
    requirement = get_object_or_404(Requirement, pk=requirement_id)
    
    if request.method == 'POST' and 'student_file' in request.FILES:
        try:
            excel_file = request.FILES['student_file']
            df = pd.read_excel(excel_file)
            
            mobile_column = request.POST.get('mobile_column', 'mobile')
            total_students = 0
            added_students = 0
            not_found_students = 0
            already_assigned = 0
            
            for index, row in df.iterrows():
                if mobile_column not in row:
                    messages.error(request, f"Column '{mobile_column}' not found in the Excel file.")
                    break
                    
                mobile = str(row[mobile_column]).strip()
                if not mobile or mobile == 'nan':
                    continue
                    
                total_students += 1
                
                try:
                    mobile = ''.join(filter(str.isdigit, mobile))
                    
                    if len(mobile) < 10:
                        not_found_students += 1
                        continue
                        
                    if len(mobile) > 10:
                        mobile = mobile[-10:]
                        
                    student = Student.objects.get(contact_number__endswith=mobile)
                    
                    if not RequirementStudent.objects.filter(requirement=requirement, student=student).exists():
                        RequirementStudent.objects.create(
                            requirement=requirement,
                            student=student
                        )
                        added_students += 1
                    else:
                        already_assigned += 1
                        
                except Student.DoesNotExist:
                    not_found_students += 1
                    continue
                except Student.MultipleObjectsReturned:
                    students = Student.objects.filter(contact_number__endswith=mobile)
                    for student in students:
                        if not RequirementStudent.objects.filter(requirement=requirement, student=student).exists():
                            RequirementStudent.objects.create(
                                requirement=requirement,
                                student=student
                            )
                            added_students += 1
                        else:
                            already_assigned += 1
            
            messages.success(request, 
                f'Import completed! '
                f'Added {added_students} students out of {total_students} from the file. '
                f'{not_found_students} students were not found in the database. '
                f'{already_assigned} students were already assigned to this requirement.'
            )
            
        except Exception as e:
            messages.error(request, f'Error processing Excel file: {str(e)}')
    
    return redirect('requirement_students', pk=requirement_id)

@login_required
def remove_requirement_student(request, requirement_id, student_id):
    requirement_student = get_object_or_404(
        RequirementStudent, 
        requirement_id=requirement_id,
        student_id=student_id
    )
    
    try:
        requirement_student.delete()
        messages.success(request, "Student successfully removed from requirement.")
    except Exception as e:
        messages.error(request, f"Error removing student: {str(e)}")
    
    return redirect('requirement_students', pk=requirement_id)

@login_required
def bulk_remove_requirement_students(request, requirement_id):
    student_ids = request.POST.getlist('student_ids', [])
    
    if not student_ids:
        messages.warning(request, "No students selected for removal.")
        return redirect('requirement_students', pk=requirement_id)
    
    try:
        deleted_count = RequirementStudent.objects.filter(
            requirement_id=requirement_id,
            student_id__in=student_ids
        ).delete()[0]
        
        messages.success(request, f"{deleted_count} students successfully removed from requirement.")
    except Exception as e:
        messages.error(request, f"Error removing students: {str(e)}")
    
    return redirect('requirement_students', pk=requirement_id)
from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from django.db.models import Q, Count, F, Prefetch
from django.contrib import messages
from .models import Requirement, Student, RequirementStudent, StudentSubjectRating, Subject

def add_students_to_requirement(request, requirement_id):
    requirement = get_object_or_404(Requirement, id=requirement_id)

    assigned_students_ids = RequirementStudent.objects.filter(
        requirement=requirement
    ).values_list('student_id', flat=True)

    students = Student.objects.exclude(
        Q(id__in=assigned_students_ids) | 
        Q(is_placed=True) | 
        Q(is_dropout=True)
    )

    if requirement.percentage_10th is not None:
        students = students.filter(tenth_percent__gte=requirement.percentage_10th)
    if requirement.percentage_12th is not None:
        students = students.filter(twelfth_percent__gte=requirement.percentage_12th)
    if requirement.percentage_degree is not None:
        students = students.filter(degree_percent__gte=requirement.percentage_degree)

    required_subjects = requirement.subjects.all()
    required_subject_ids = list(required_subjects.values_list('id', flat=True))
    has_other_subject = required_subjects.filter(name='other').exists()

    if required_subjects.exists() and not has_other_subject:
        students = students.annotate(
            matching_subjects_count=Count(
                'subject_ratings',
                filter=Q(subject_ratings__subject__in=required_subjects)
            )
        ).filter(matching_subjects_count=len(required_subject_ids))

        students = students.prefetch_related(
            Prefetch(
                'subject_ratings',
                queryset=StudentSubjectRating.objects.filter(
                    subject__in=required_subjects
                ).select_related('subject'),
                to_attr='prefetched_required_ratings'
            )
        )
    elif has_other_subject:
        students = students.prefetch_related(
            Prefetch(
                'subject_ratings',
                queryset=StudentSubjectRating.objects.all().select_related('subject'),
                to_attr='all_subject_ratings'
            )
        )

    query = request.GET.get('q')
    if query:
        students = students.filter(
            Q(name__icontains=query) |
            Q(contact_number__icontains=query) |
            Q(degree__icontains=query) |
            Q(stream__icontains=query)
        )

    stream_filters = request.GET.getlist('stream')
    if stream_filters:
        stream_query = Q()
        for stream in stream_filters:
            stream_query |= Q(stream__iexact=stream)
        students = students.filter(stream_query)

    degree_filters = request.GET.getlist('degree')
    if degree_filters:
        degree_query = Q()
        for degree in degree_filters:
            degree_query |= Q(degree__iexact=degree)
        students = students.filter(degree_query)

    start_year = request.GET.get('start_year')
    end_year = request.GET.get('end_year')
    if start_year and end_year and start_year.isdigit() and end_year.isdigit():
        students = students.filter(yop__range=(start_year, end_year))

    tenth = request.GET.get('tenth')
    if tenth and tenth.replace('.', '', 1).isdigit():
        students = students.filter(tenth_percent__gte=float(tenth))

    twelfth = request.GET.get('twelfth')
    if twelfth and twelfth.replace('.', '', 1).isdigit():
        students = students.filter(twelfth_percent__gte=float(twelfth))

    degree_percentage = request.GET.get('degree_percentage')
    if degree_percentage and degree_percentage.replace('.', '', 1).isdigit():
        students = students.filter(degree_percent__gte=float(degree_percentage))

    gender_filter = request.GET.get('gender')
    if gender_filter:
        students = students.filter(gender=gender_filter)

    data_type_filter = request.GET.get('data_type')
    if data_type_filter:
        students = students.filter(type_of_data=data_type_filter)

    students = students.order_by(
        F('scheduled_requirements').asc(nulls_first=True),
        F('total_requirements').asc(nulls_first=True),
        'name'
    ).distinct()

    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_students')
        if selected_ids:
            count = 0
            for student_id in selected_ids:
                student = Student.objects.get(id=student_id)
                if not RequirementStudent.objects.filter(
                    student=student,
                    requirement=requirement
                ).exists():
                    RequirementStudent.objects.create(
                        student=student,
                        requirement=requirement,
                        status='pending'
                    )
                    count += 1

            if count > 0:
                messages.success(request, f'Successfully assigned {count} student(s) to the requirement.')
            else:
                messages.info(request, 'No new students were assigned.')
            return redirect('student_data:add_students_to_requirement', requirement_id=requirement.id)

    paginator = Paginator(students, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    enhanced_students = []
    for student in page_obj:
        progress_percentage = 0
        if student.total_requirements and student.total_requirements > 0:
            progress_percentage = (student.scheduled_requirements / student.total_requirements) * 100

        subject_ratings = []
        if required_subjects.exists() and not has_other_subject:
            ratings_dict = {
                rating.subject.name: rating for rating in getattr(student, 'prefetched_required_ratings', [])
            }
            for subject in required_subjects:
                rating = ratings_dict.get(subject.name)
                subject_ratings.append({
                    'subject': subject.get_name_display(),
                    'rating': rating.get_rating_display() if rating else 'Not Rated',
                    'remarks': rating.remarks if rating else ''
                })
        # else: skip showing any ratings if "other" is included

        enhanced_students.append({
            'student': student,
            'progress_percentage': progress_percentage,
            'subject_ratings': subject_ratings,
            'is_assigned': student.id in assigned_students_ids,
        })

    def get_eligible_uppercase_values(field_name):
        return sorted({
            value.upper() for value in 
            students.exclude(**{f'{field_name}__isnull': True})
                    .exclude(**{f'{field_name}__exact': ''})
                    .values_list(field_name, flat=True)
            if value.strip()
        })

    context = {
        'requirement': requirement,
        'enhanced_students': enhanced_students,
        'page_obj': page_obj,
        'required_subjects': required_subjects,
        'has_other_subject': has_other_subject,
        'unique_streams': get_eligible_uppercase_values('stream'),
        'unique_degrees': get_eligible_uppercase_values('degree'),
        'unique_data_types': Student.objects.values_list('type_of_data', flat=True)
                          .distinct().order_by('type_of_data'),
        'selected_data_type': data_type_filter,
        'selected_streams': [s.upper() for s in stream_filters],
        'degree_filters': [d.upper() for d in degree_filters],
        'start_year': start_year,
        'end_year': end_year,
        'tenth': tenth,
        'twelfth': twelfth,
        'degree_percentage': degree_percentage,
        'query': query,
        'gender_filter': gender_filter,
        'total_available': students.count(),
        'total_filtered': page_obj.paginator.count,
    }

    return render(request, 'add_student_to_requirement.html', context)

from django.utils import timezone
from datetime import timedelta
from .models import Requirement, Student

@login_required
def home_dashboard(request):
    today = timezone.now().date()
    tomorrow = today + timedelta(days=1)
    
    context = {
        'total_students': Student.objects.count(),
        'total_requirements': Requirement.objects.count(),
        'scheduled_today': Requirement.objects.filter(
            schedule_status='scheduled',
            schedule_date=today  # No __date needed for DateField
        ).count(),
        'scheduled_tomorrow': Requirement.objects.filter(
            schedule_status='scheduled',
            schedule_date=tomorrow  # Same here
        ).count(),
        'latest_requirements': Requirement.objects.all().order_by('-created_at')[:5],
        'todays_requirements': Requirement.objects.filter(
            schedule_status='scheduled',
            schedule_date=today
        ).order_by('schedule_time'),
        'recent_students': Student.objects.all().order_by('-created_at')[:5],
        'requirement_stats': Requirement.objects.values('schedule_status').annotate(
            count=Count('id')
        ),
    }
    
    return render(request, 'home_dashboard.html', context)

# Analytics views
@login_required
def student_analytics(request):
    students = Student.objects.all().values(
        'name', 'degree', 'stream', 'tenth_percent', 
        'twelfth_percent', 'degree_percent', 'type_of_data'
    )
    df = pd.DataFrame(list(students))

    type_distribution = df['type_of_data'].value_counts()
    type_chart = px.pie(
        values=type_distribution.values,
        names=type_distribution.index,
        title='Distribution of Students by Type',
        template='plotly_dark'
    )

    performance_fig = go.Figure()
    performance_fig.add_trace(go.Box(y=df['tenth_percent'], name='10th'))
    performance_fig.add_trace(go.Box(y=df['twelfth_percent'], name='12th'))
    performance_fig.add_trace(go.Box(y=df['degree_percent'], name='Degree'))
    performance_fig.update_layout(
        title='Performance Distribution Across Education Levels',
        yaxis_title='Percentage',
        template='plotly_dark'
    )

    stream_avg = df.groupby('stream')[['tenth_percent', 'twelfth_percent', 'degree_percent']].mean()
    stream_fig = px.bar(
        stream_avg,
        title='Average Performance by Stream',
        barmode='group',
        template='plotly_dark'
    )

    requirements_data = RequirementStudent.objects.values(
        'requirement__company_name'
    ).annotate(
        count=Count('id'),
        scheduled=Count('id', filter=Q(requirement__is_scheduled=True))
    )
    req_df = pd.DataFrame(list(requirements_data))
    
    if not req_df.empty:
        req_fig = px.bar(
            req_df,
            x='requirement__company_name',
            y=['count', 'scheduled'],
            title='Requirements Status by Company',
            template='plotly_dark'
        )
    else:
        req_fig = go.Figure()

    charts = {
        'type_distribution': type_chart.to_json(),
        'performance_analysis': performance_fig.to_json(),
        'stream_analysis': stream_fig.to_json(),
        'requirements_analysis': req_fig.to_json(),
    }

    summary_stats = {
        'total_students': len(df),
        'avg_degree_percent': df['degree_percent'].mean(),
        'top_stream': df['stream'].mode().iloc[0],
        'type_counts': df['type_of_data'].value_counts().to_dict()
    }

    context = {
        'charts': charts,
        'summary_stats': summary_stats,
    }

    return render(request, 'analytics/student_analytics.html', context)

@login_required
def performance_trends(request):
    students = Student.objects.all().values(
        'name', 'tenth_percent', 'twelfth_percent', 
        'degree_percent', 'type_of_data', 'stream'
    )
    df = pd.DataFrame(list(students))

    fig = make_subplots(rows=2, cols=2)

    fig.add_trace(
        go.Scatter(
            x=['10th', '12th', 'Degree'],
            y=[df['tenth_percent'].mean(), 
               df['twelfth_percent'].mean(), 
               df['degree_percent'].mean()],
            name='Average Progression'
        ),
        row=1, col=1
    )

    for student_type in df['type_of_data'].unique():
        type_df = df[df['type_of_data'] == student_type]
        fig.add_trace(
            go.Box(
                y=type_df['degree_percent'],
                name=student_type,
                showlegend=True
            ),
            row=1, col=2
        )

    stream_perf = df.groupby('stream')['degree_percent'].mean().sort_values(ascending=False)
    fig.add_trace(
        go.Bar(
            x=stream_perf.index,
            y=stream_perf.values,
            name='Stream Performance'
        ),
        row=2, col=1
    )

    fig.add_trace(
        go.Scatter(
            x=df['twelfth_percent'],
            y=df['degree_percent'],
            mode='markers',
            name='12th vs Degree'
        ),
        row=2, col=2
    )

    fig.update_layout(height=800, title_text="Performance Analysis Dashboard")
    
    context = {
        'performance_dashboard': fig.to_json(),
    }

    return render(request, 'analytics/performance_trends.html', context)


from django.shortcuts import render
from django.utils import timezone
from .models import Requirement

def todays_requirements_view(request):
    today = timezone.now().date()
    requirements = Requirement.objects.filter(
        schedule_date=today,
        is_scheduled=True
    ).order_by('schedule_time')

    context = {
        'requirements': requirements
    }
    return render(request, 'todays_requirements.html', context)

from django.utils import timezone
from django.shortcuts import render
from .models import Requirement, RequirementStudent

from django.utils import timezone
from django.shortcuts import render
from django.contrib import messages
from .models import Requirement, RequirementStudent

def students_attending_today(request):
    today = timezone.localdate()
    print(f"DEBUG: Checking attendance for {today}")  # Debug output
    
    # Get requirements scheduled for today
    requirements_today = Requirement.objects.filter(
        schedule_date=today,
        schedule_status='scheduled'
    ).order_by('schedule_time').prefetch_related('requirementstudent_set__student')

    if not requirements_today.exists():
        debug_msg = (
            f"No scheduled requirements found for {today}. "
            "Please check:\n"
            f"- Requirements with schedule_date = {today}\n"
            "- Requirements with schedule_status = 'scheduled'"
        )
        print(debug_msg)  # Console debug
        messages.warning(request, debug_msg)
        return render(request, 'students_attending_today.html', {
            'attendance_data': [],
            'today': today.strftime("%B %d, %Y"),
            'debug_info': {
                'requirements_count': 0,
                'total_students': 0,
                'error': debug_msg
            }
        })

    attendance_data = []
    total_students = 0

    for requirement in requirements_today:
        students = []
        requirement_students = requirement.requirementstudent_set.all()
        
        if not requirement_students.exists():
            print(f"DEBUG: Requirement {requirement.id} has no students assigned")

        for rs in requirement_students:
            students.append({
                'id': rs.student.id,
                'name': rs.student.name,
                'status': rs.get_status_display(),
                'feedback': rs.feedback,
                'requirement_student_id': rs.id  # For debugging
            })

        student_count = len(students)
        total_students += student_count
        
        attendance_data.append({
            'company': requirement.company_name,
            'time': requirement.schedule_time.strftime("%H:%M") if requirement.schedule_time else "Not specified",
            'students': students,
            'total_students': student_count,
            'requirement_id': requirement.id,
            'schedule_date': requirement.schedule_date,
            'schedule_status': requirement.schedule_status,
            'requirement_date':requirement.requirement_date,
        })

    context = {
        'attendance_data': attendance_data,
        'today': today.strftime("%B %d, %Y"),
        'debug_info': {
            'requirements_count': requirements_today.count(),
            'total_students': total_students,
            'requirements_list': [r.id for r in requirements_today]
        }
     } #
    return render(request, 'students_attending_today.html', context)

import io
import zipfile
import pandas as pd
from django.http import HttpResponse
from django.utils import timezone
from .models import RequirementStudent, Requirement, Student
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill


def export_today_scheduled_requirements(request):
    today = timezone.now().date()
    requirements = Requirement.objects.filter(schedule_date=today, schedule_status='scheduled')
    
    if not requirements.exists():
        return HttpResponse("No scheduled requirements for today.", status=404)

    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for req in requirements:
            req_students = RequirementStudent.objects.filter(
                requirement=req
            ).select_related('student').order_by('student__name')
            
            student_data = []
            for req_student in req_students:
                student = req_student.student
                student_data.append({
                    'Student Name': student.name,
                    'Contact Number': student.contact_number,
                    'Gender': student.gender,
                    'Degree': student.degree,
                    'Stream': student.stream,
                    'Year of Passing': student.yop,
                    '10th %': student.tenth_percent,
                    '12th %': student.twelfth_percent,
                    'Degree %': student.degree_percent,
                    'Type of Data': student.type_of_data,
                    'Status': req_student.status,
                    'Feedback': req_student.feedback or '',
                    'Is Placed': "Yes" if student.is_placed else "No"
                })

            student_df = pd.DataFrame(student_data)

            requirement_details = {
                'Requirement Field': [
                    'Company Name',
                    'Company Code',
                    'Schedule Date',
                    'Schedule Time',
                    'Requirement Date',
                    'Description',
                    'Schedule Status',
                    'Escalation'
                ],
                'Value': [
                    req.company_name,
                    req.company_code,
                    str(req.schedule_date),
                    str(req.schedule_time),
                    str(req.requirement_date),
                    req.description,
                    req.schedule_status,
                    req.escalation
                ]
            }
            details_df = pd.DataFrame(requirement_details)

            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                details_df.to_excel(writer, index=False, sheet_name='Requirement Details')
                student_df.to_excel(writer, index=False, sheet_name='Students')

                workbook = writer.book
                details_sheet = writer.sheets['Requirement Details']
                students_sheet = writer.sheets['Students']

                # Format both sheets
                header_font = Font(bold=True)
                header_fill = PatternFill("solid", fgColor="DDDDDD")

                # Format Requirement Details Sheet
                for col in range(1, 3):  # A and B columns
                    col_letter = get_column_letter(col)
                    details_sheet.column_dimensions[col_letter].width = 30
                    cell = details_sheet[f"{col_letter}1"]
                    cell.font = header_font
                    cell.fill = header_fill

                # Format Students Sheet
                for col_idx, column in enumerate(student_df.columns, 1):
                    max_length = max(
                        student_df[column].astype(str).map(len).max(),
                        len(column)
                    )
                    col_letter = get_column_letter(col_idx)
                    students_sheet.column_dimensions[col_letter].width = min(max_length + 2, 30)
                    header_cell = students_sheet[f"{col_letter}1"]
                    header_cell.font = header_font
                    header_cell.fill = header_fill

            excel_buffer.seek(0)
            filename = f"{req.company_name.replace(' ', '_')}_{req.id}_{today}.xlsx"
            zip_file.writestr(filename, excel_buffer.read())

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="scheduled_requirements_{today}.zip"'
    return response

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Requirement, Student, RequirementStudent

def remove_student_from_requirement(request, requirement_id, student_id):
    if request.method == 'POST':
        requirement = get_object_or_404(Requirement, id=requirement_id)
        student = get_object_or_404(Student, id=student_id)

        try:
            link = RequirementStudent.objects.get(requirement=requirement, student=student)
            link.delete()
            student.update_requirement_counts()
            messages.success(request, f'{student.name} removed successfully from the requirement.')
        except RequirementStudent.DoesNotExist:
            messages.error(request, 'Student is not part of this requirement.')

        return redirect('student_data:requirement_detail', pk=requirement.id)

    messages.error(request, 'Invalid request method.')
    return redirect('student_data:requirement_detail', requirement_id=requirement.id)




from django.shortcuts import render
from django.views.decorators.csrf import requires_csrf_token
from django.template import RequestContext

@requires_csrf_token
def custom_page_not_found_view(request, exception):
    return render(request, "404.html", status=404)

# Temporary view in views.py
def test_404_template(request):
    return render(request, "404.html")


# views.py
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.views.generic import DeleteView
from .models import Student

class StudentDeleteView(LoginRequiredMixin, DeleteView):
    model = Student
    success_url = reverse_lazy('student_data:student_list')  # Update with your actual URL name
    template_name = 'student_confirm_delete.html'
    context_object_name = 'student'

    def get_queryset(self):
        """Ensure users can only delete their own students if needed"""
        qs = super().get_queryset()
        # Add any additional filtering logic here if needed
        return qs
    
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Requirement
import openpyxl
from datetime import datetime as dt  # Import with alias to avoid confusion
from datetime import date

def delete_all_requirements(request):
    if request.method == 'POST':
        Requirement.objects.all().delete()
        messages.success(request, "✅ All requirements deleted and students updated.")
    else:
        messages.error(request, "❌ Invalid request method.")
    return redirect('student_data:requirement_list')

def parse_excel_date(cell_value, row_index, field_name, failed_rows):
    """
    Parse dates from Excel safely without swapping day/month
    """
    if cell_value is None or (isinstance(cell_value, str) and not cell_value.strip()):
        return None

    # Handle datetime/date objects from Excel
    if isinstance(cell_value, (dt, date)):
        try:
            return cell_value if isinstance(cell_value, date) else cell_value.date()
        except Exception as e:
            failed_rows.append(f"Row {row_index}: Error processing date object - {str(e)}")
            return None

    # Handle string dates
    if isinstance(cell_value, str):
        cell_value = cell_value.strip()
        for fmt in ("%d-%m-%Y", "%m-%d-%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return dt.strptime(cell_value, fmt).date()
            except ValueError:
                continue

        failed_rows.append(
            f"Row {row_index}: Invalid {field_name} format ('{cell_value}') - must be DD-MM-YYYY, MM-DD-YYYY, DD/MM/YYYY or MM/DD/YYYY"
        )
        return None

    failed_rows.append(f"Row {row_index}: Invalid {field_name} type")
    return None

def upload_requirements_view(request):
    message = None
    success_count = 0
    failed_rows = []

    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active

            for i, row in enumerate(sheet.iter_rows(min_row=3), start=3):
                try:
                    company_name = row[1].value
                    company_code = row[2].value
                    requirement_date_raw = row[3].value
                    schedule_date_raw = row[4].value

                    if not company_code or not company_name:
                        failed_rows.append(f"Row {i}: Missing company name/code")
                        continue

                    requirement_date = parse_excel_date(
                        requirement_date_raw, i, "requirement date", failed_rows
                    )
                    if requirement_date is None:
                        continue

                    schedule_date = None
                    if schedule_date_raw is not None and str(schedule_date_raw).strip():
                        schedule_date = parse_excel_date(
                            schedule_date_raw, i, "schedule date", failed_rows
                        )
                        if schedule_date is None:
                            continue

                    Requirement.objects.update_or_create(
                        company_code=company_code,
                        defaults={
                            'company_name': company_name,
                            'requirement_date': requirement_date,
                            'schedule_date': schedule_date,
                            'is_scheduled': schedule_date is not None,
                            'schedule_status': 'scheduled' if schedule_date else 'not_scheduled'
                        }
                    )
                    success_count += 1

                except Exception as e:
                    failed_rows.append(f"Row {i}: Error processing row - {str(e)}")

            message = f"Successfully uploaded {success_count} requirement(s)."
            if failed_rows:
                message += f" Failed to process {len(failed_rows)} row(s)."

        except Exception as e:
            failed_rows.append(f"Error processing file: {str(e)}")
            message = "Failed to process the uploaded file."

    return render(request, 'upload_requirements.html', {
        'message': message,
        'errors': failed_rows
    })
# views.py (add this new view)
import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Requirement, Student, RequirementStudent

# views.py\
@login_required
def map_students_to_requirement_view(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active
        success_count = 0
        errors = []

        # Find column indices by header names
        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        column_indices = {
            'mobile_number': None,
            'company_code': None
        }

        for idx, cell in enumerate(header_row):
            header_name = str(cell.value).strip().lower()
            if 'mobile' in header_name:
                column_indices['mobile_number'] = idx
            elif 'company' in header_name and 'code' in header_name:
                column_indices['company_code'] = idx

        # Validate headers
        if None in column_indices.values():
            errors.append("Missing required columns: Mobile number and/or Company code")
            messages.error(request, "Invalid file format")
            return render(request, 'map_students.html', {'errors': errors})

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            try:
                # Get values from correct columns
                mobile = str(row[column_indices['mobile_number']].value).strip()
                company_code = str(row[column_indices['company_code']].value).strip()

                if not mobile or not company_code:
                    errors.append(f"Row {row_idx}: Missing required data")
                    continue

                # Case-sensitive company code lookup
                requirement = Requirement.objects.get(company_code__exact=company_code)
                student = Student.objects.get(contact_number=mobile)

                # Create mapping if not exists
                _, created = RequirementStudent.objects.get_or_create(
                    requirement=requirement,
                    student=student,
                    defaults={'status': 'pending'}
                )

                if created:
                    success_count += 1
                else:
                    errors.append(f"Row {row_idx}: Mapping already exists for {mobile} - {company_code}")

            except Requirement.DoesNotExist:
                errors.append(f"Row {row_idx}: Requirement with code '{company_code}' not found")
            except Student.DoesNotExist:
                errors.append(f"Row {row_idx}: Student with mobile '{mobile}' not found")
            except Exception as e:
                errors.append(f"Row {row_idx}: Error - {str(e)}")

        if success_count > 0:
            messages.success(request, f"Successfully mapped {success_count} students!")
        if errors:
            messages.error(request, f"Completed with {len(errors)} errors")

        return render(request, 'map_students.html', {'errors': errors})

    return render(request, 'map_students.html')

from django.http import HttpResponse
from openpyxl import Workbook
from django.views.decorators.http import require_GET
from django.utils.timezone import now

@require_GET
def download_sample_excel_map(request):
    """Generate and return a sample Excel template for student-requirement mapping"""
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping Template"
    
    # Create headers with exact expected naming
    headers = ["Mobile number", "Company code"]
    ws.append(headers)
    
    # Add example data rows
    sample_data = [
        ("9876543210", "COMPANY_123"),
        ("9123456789", "COMPANY_456"),
        ("9555512345", "CORP_2024"),
    ]
    for data_row in sample_data:
        ws.append(data_row)
    
    # Format response
    timestamp = now().strftime("%Y%m%d_%H%M")
    filename = f"student_mapping_template_{timestamp}.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )
    
    # Save workbook directly to response
    wb.save(response)
    return response


from django.shortcuts import render, redirect
from django.contrib import messages
from django.urls import reverse
from django.utils.http import urlencode
from .models import Student, Requirement, RequirementStudent

@login_required
def combined_view(request):
    student = None
    requirement = None
    req_students = None
    mobile_number = request.GET.get("mobile_number")
    company_code = request.GET.get("company_code")

    if request.method == "POST" and 'req_student_id' in request.POST:
        req_student_id = request.POST.get("req_student_id")
        new_status = request.POST.get("status")
        mobile_number = request.POST.get("mobile_number")
        company_code = request.POST.get("company_code")

        try:
            req_student = RequirementStudent.objects.get(id=req_student_id)
            if new_status in dict(RequirementStudent.STATUS_CHOICES):
                old_status = req_student.status
                req_student.status = new_status
                req_student.save()

                # Update student placement status if status changed to/from 'selected'
                if new_status == 'selected' or old_status == 'selected':
                    student = req_student.student
                    student.update_placement_status()
                    student.save()

                messages.success(request, "Status updated successfully.")
            else:
                messages.error(request, "Invalid status.")
        except RequirementStudent.DoesNotExist:
            messages.error(request, "Mapping not found.")

        redirect_url = reverse('student_data:update_status')
        params = {}
        if mobile_number:
            params['mobile_number'] = mobile_number
        elif company_code:
            params['company_code'] = company_code
        if params:
            redirect_url += '?' + urlencode(params)
        return redirect(redirect_url)

    # Handle GET requests
    if mobile_number:
        try:
            student = Student.objects.get(contact_number=mobile_number)
            req_students = RequirementStudent.objects.filter(student=student).select_related('requirement')
        except Student.DoesNotExist:
            messages.error(request, "Student not found.")

    elif company_code:
        try:
            requirement = Requirement.objects.get(company_code=company_code)
            req_students = RequirementStudent.objects.filter(requirement=requirement).select_related('student')
        except Requirement.DoesNotExist:
            messages.error(request, "Company code not found.")

    context = {
        "student": student,
        "requirement": requirement,
        "req_students": req_students,
        "STATUS_CHOICES": RequirementStudent.STATUS_CHOICES,
        "mobile_number": mobile_number,
        "company_code": company_code,
    }
    return render(request, "combined_template.html", context)


from django.core.paginator import Paginator
from django.shortcuts import render

from django.shortcuts import render
from .models import RequirementStudent
@login_required
def placed_students_view(request):
    placed_students = RequirementStudent.objects.filter(
        status='selected',
        student__is_placed=True,
        student__is_dropout=False
    ).exclude(
        student__placed_outside__isnull=False
    ).select_related('student').distinct().order_by('-created_at')

    # Get unique student types and their counts
    student_types = Student.objects.filter(
        is_placed=True,
        is_dropout=False
    ).exclude(placed_outside__isnull=False).values_list('type_of_data', flat=True).distinct()

    type_counts = {}
    for stype in student_types:
        type_counts[stype] = placed_students.filter(student__type_of_data=stype).count()

    return render(request, 'placed_students.html', {
        'placed_students': placed_students,
        'total_count': placed_students.count(),
        'type_counts': type_counts,
    })

from django.http import HttpResponse
import csv

def export_placed_students_excel(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="placed_students.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Student Name', 'Company', 'Requirements', 'Scheduled Requirements', 'Schedule Date'])
    
    for entry in RequirementStudent.objects.filter(status='selected'):
        writer.writerow([
            entry.student.name,
            entry.requirement.company_name,
            entry.student.total_requirements,
            entry.student.scheduled_requirements,
            entry.requirement.schedule_date or 'Not Scheduled'
        ])
    
    return response
def delete_requirement(request, pk):
    requirement = get_object_or_404(Requirement, pk=pk)
    
    if request.method == 'POST':
        requirement.delete()
        messages.success(request, f'Requirement "{requirement.company_name}" deleted successfully.')
        return redirect('student_data:requirement_list')
    
    context = {
        'requirement': requirement,
    }
    return render(request, 'delete_requirement.html', context)


import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from .models import Student, Requirement, RequirementStudent
from io import BytesIO
from django.core.exceptions import ObjectDoesNotExist
@login_required
def update_selected_students(request):
    context = {
        'existing_companies': Requirement.objects.exclude(
            company_code__isnull=True
        ).exclude(
            company_code__exact=''
        ).values_list('company_code', flat=True).distinct().order_by('company_code')
    }

    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('excel_file')
            mobile_column = request.POST.get('mobile_column', 'mobile_number').strip()
            company_column = request.POST.get('company_column', 'company_code').strip()

            if not excel_file:
                raise ValueError("Please upload an Excel file")

            # Read file
            try:
                if excel_file.name.endswith('.xlsx'):
                    df = pd.read_excel(excel_file)
                elif excel_file.name.endswith('.csv'):
                    df = pd.read_csv(excel_file)
                else:
                    raise ValueError("Unsupported file format")
            except Exception as e:
                raise ValueError(f"Error reading file: {str(e)}")

            # Validate columns
            for col in [mobile_column, company_column]:
                if col not in df.columns:
                    raise ValueError(
                        f"Column '{col}' not found in file. Available columns: {', '.join(df.columns)}"
                    )

            df = df.dropna(subset=[mobile_column, company_column])
            df[mobile_column] = df[mobile_column].astype(str).str.replace(r'\D', '', regex=True)
            df[company_column] = df[company_column].astype(str).str.strip()

            with transaction.atomic():
                results = {
                    'companies_processed': set(),
                    'students_selected': 0,
                    'records_created': 0,
                    'records_updated': 0,
                    'students_rejected': 0,
                    'numbers_not_found': 0,
                    'invalid_companies': set(),
                    'processed_count': len(df)
                }

                for _, row in df.iterrows():
                    mobile = row[mobile_column]
                    company_code = row[company_column]

                    if not mobile or not company_code:
                        continue

                    try:
                        requirement = Requirement.objects.get(company_code=company_code)
                        results['companies_processed'].add(company_code)
                    except ObjectDoesNotExist:
                        results['invalid_companies'].add(company_code)
                        continue

                    try:
                        student = Student.objects.get(contact_number=mobile)
                    except ObjectDoesNotExist:
                        results['numbers_not_found'] += 1
                        continue

                    obj, created = RequirementStudent.objects.update_or_create(
                        student=student,
                        requirement=requirement,
                        defaults={'status': 'selected'}
                    )

                    if created:
                        results['records_created'] += 1
                    elif obj.status != 'selected':
                        obj.status = 'selected'
                        obj.save()
                        results['records_updated'] += 1

                    # Update student's placement status
                    if not student.is_placed:
                        student.is_placed = True
                        student.save()

                    results['students_selected'] += 1
                    student.update_requirement_counts()

                # Mark others as rejected
                for company_code in results['companies_processed']:
                    try:
                        requirement = Requirement.objects.get(company_code=company_code)
                        selected_students = Student.objects.filter(
                            requirementstudent__requirement=requirement,
                            requirementstudent__status='selected'
                        )
                        rejected = RequirementStudent.objects.filter(
                            requirement=requirement,
                            status='pending'
                        ).exclude(student__in=selected_students).update(status='rejected')
                        results['students_rejected'] += rejected
                    except:
                        continue

                msg = [
                    f"<strong>Processed {results['processed_count']} rows</strong>",
                    f"Companies processed: {len(results['companies_processed'])}",
                    f"Students selected: {results['students_selected']}",
                    f"New records created: {results['records_created']}",
                    f"Existing records updated: {results['records_updated']}",
                    f"Students marked as rejected: {results['students_rejected']}",
                    f"Phone numbers not found: {results['numbers_not_found']}",
                ]
                if results['invalid_companies']:
                    sample = ", ".join(list(results['invalid_companies'])[:3])
                    msg.append(f"Invalid company codes ({len(results['invalid_companies'])}): {sample}{'...' if len(results['invalid_companies']) > 3 else ''}")

                messages.success(request, "<br>".join(msg))
                context.update(results)

        except Exception as e:
            messages.error(request, f"<strong>Error:</strong> {str(e)}")

    return render(request, 'update_selected_students.html', context)


import io
import pandas as pd
from django.http import HttpResponse

def download_student_selection_template(request):
    # Sample DataFrame
    df = pd.DataFrame({
        'mobile_number': ['9876543210', '9123456780'],
        'company_code': ['ABC123', 'XYZ456']
    })

    # Save to in-memory file
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Template')

    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=student_selection_template.xlsx'
    return response

from django.contrib import messages
from django.urls import reverse_lazy
from django.views.generic.edit import CreateView
from .models import Student

class StudentCreateView(CreateView):
    model = Student
    fields = [
        'name', 
        'contact_number', 
        'degree', 
        'stream', 
        'yop',
        'tenth_percent', 
        'twelfth_percent', 
        'degree_percent',
        'gender', 
        'type_of_data',
        'is_dropout',
        'dropout_date',
        'dropout_reason',
        'overall_technical_rating',
    ]
    template_name = 'student_add.html'
    success_url = reverse_lazy('student_data:add')
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Add placeholders or help text if needed
        form.fields['name'].help_text = "Enter full name"
        form.fields['contact_number'].help_text = "Include country code if international"
        form.fields['yop'].label = "Year of Passing"
        form.fields['tenth_percent'].label = "10th Percentage"
        form.fields['twelfth_percent'].label = "12th Percentage"
        form.fields['degree_percent'].label = "Degree Percentage"
        
        # Make dropout fields conditional
        if not self.object or not self.object.is_dropout:
            form.fields['dropout_date'].widget.attrs['disabled'] = True
            form.fields['dropout_reason'].widget.attrs['disabled'] = True
            
        return form
    
    def form_valid(self, form):
        # Clean and process data before saving
        instance = form.save(commit=False)
        
        # Handle dropout fields logic
        if not instance.is_dropout:
            instance.dropout_date = None
            instance.dropout_reason = None
            
        instance.save()
        form.save_m2m()  # In case there are many-to-many fields added later
        
        messages.success(self.request, 'Student added successfully!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Add New Student"
        return context


from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Protection

def download_template(request):
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Add headers
    headers = [
        "S.No", 
        "Company Name", 
        "Company Code", 
        "Requirement Date (DD-MM-YYYY)", 
        "Schedule Date (DD-MM-YYYY)"
    ]
    ws.append(headers)
    
    # Add sample data
    sample_data = [
        ["1", "ABC Corp", "ABC001", "15-01-2025", "20-01-2025"],
        ["2", "XYZ Ltd", "XYZ002", "10-02-2025", ""],
    ]
    for row in sample_data:
        ws.append(row)
    
    # Format date columns as text to prevent Excel auto-formatting
    for cell in ws['D'] + ws['E']:
        cell.protection = Protection(locked=False)
        if cell.row > 1:  # Skip header
            cell.number_format = '@'  # Text format
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="requirements_template.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    return response

from django.views.generic import UpdateView
class RaiseEscalationView(UpdateView):
    model = Requirement
    fields = ['escalation']
    template_name = 'raise_escalation.html'
    success_url = reverse_lazy('student_data:requirement_list')  # Change this to your list view URL

    def form_valid(self, form):
        requirement = form.save(commit=False)
        requirement.escalation_raised_at = timezone.now()
        requirement.save()
        messages.success(self.request, 'Escalation raised successfully!')
        return super().form_valid(form)



from django.views.generic import TemplateView
from django.db.models import Count, Q, F
from django.db.models.functions import TruncMonth
from django.utils.timezone import now
from django.http import HttpResponse
import datetime
import openpyxl
from .models import Requirement, RequirementStudent

class MonthlyReportsView(TemplateView):
    template_name = 'monthly_reports.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_month = now().date().replace(day=1)
        context['current_month'] = current_month

        # Get all months with requirements for dropdown
        months = Requirement.objects.exclude(schedule_date__isnull=True).dates('schedule_date', 'month', order='DESC')
        context['months'] = months
        context['monthly_summary'] = self.get_monthly_summary_data()

        selected_month = self.request.GET.get('month')
        if selected_month:
            try:
                from datetime import datetime
                month_date = datetime.strptime(selected_month, '%Y-%m').date()
                context['selected_month'] = month_date
                month_details = self.get_month_details(month_date)
                context.update(month_details)
                
                # Add not scheduled and pending requirements to context
                context['not_scheduled_list'] = month_details.get('not_scheduled_list', [])
                context['pending_requirements'] = month_details.get('pending_requirements', [])
                
                cross_reqs = self.get_cross_month_requirements(month_date)
                context['cross_month_requirements'] = [
                    {
                        'id': req.id,
                        'company_name': req.company_name,
                        'company_code': req.company_code,
                        'requirement_date': req.requirement_date,
                        'schedule_date': req.schedule_date,
                        'schedule_status': req.schedule_status,
                        'scheduled_details': req.scheduled_details,
                        'escalation': req.escalation or '',
                    }
                    for req in cross_reqs
                ]
            except ValueError:
                pass
        else:
            # Show current month by default
            context.update(self.get_month_details(current_month))
            context['cross_month_requirements'] = []

        return context

    def get_monthly_summary_data(self):
        months = Requirement.objects.annotate(
            month=TruncMonth('requirement_date')
        ).values('month').annotate(
            total=Count('id'),
            scheduled=Count('id', filter=Q(schedule_date__isnull=False)),
            not_scheduled=Count('id', filter=Q(schedule_date__isnull=True)),
            escalated=Count('id', filter=Q(escalation__isnull=False) & ~Q(escalation='')),
            pending=Count('id', filter=Q(schedule_date__isnull=False) & Q(scheduled_details__result__isnull=True)),
        ).order_by('-month')

        # Completed Rounds
        completed_reqs = Requirement.objects.annotate(
            month=TruncMonth('requirement_date')
        ).values('month').annotate(
            completed_rounds=Count('id', filter=Q(
                Q(scheduled_details__result__in=['selected', 'no_selects']) |
                Q(requirementstudent__status='selected')
            ), distinct=True)
        )
        completed_dict = {item['month']: item['completed_rounds'] for item in completed_reqs}

        # Placements
        placement_counts = RequirementStudent.objects.filter(
            status='selected'
        ).annotate(
            month=TruncMonth('requirement__requirement_date')
        ).values('month').annotate(
            placed=Count('student', distinct=True)
        )
        placement_dict = {item['month']: item['placed'] for item in placement_counts}

        summaries = []
        for m in months:
            m['completed_rounds'] = completed_dict.get(m['month'], 0)
            m['placed'] = placement_dict.get(m['month'], 0)
            summaries.append(m)

        return summaries

    def get_month_details(self, month_date):
        details = {}

        # Requirements posted in that month
        month_reqs = Requirement.objects.filter(
            requirement_date__year=month_date.year,
            requirement_date__month=month_date.month
        ).select_related('scheduled_details')

        total_reqs = month_reqs.count()
        scheduled_reqs = month_reqs.filter(schedule_date__isnull=False)
        not_scheduled_reqs = month_reqs.filter(schedule_date__isnull=True)
        escalated_reqs = month_reqs.exclude(escalation__isnull=True).exclude(escalation='')

        # Metrics
        scheduled_companies = scheduled_reqs.count()
        not_scheduled_companies = not_scheduled_reqs.count()
        escalated_companies = escalated_reqs.count()

        # Pending requirements (scheduled but no result updated)
        pending_reqs = scheduled_reqs.filter(scheduled_details__result__isnull=True)
        pending_companies = pending_reqs.count()

        # Completed rounds
        all_rounds_completed = scheduled_reqs.filter(
            Q(scheduled_details__result__in=['selected', 'no_selects']) |
            Q(requirementstudent__status='selected')
        ).distinct().count()

        # Placements
        placements = RequirementStudent.objects.filter(
            status='selected',
            requirement__requirement_date__year=month_date.year,
            requirement__requirement_date__month=month_date.month
        ).values('student').distinct().count()

        # Prepare not scheduled list
        not_scheduled_list = []
        for req in not_scheduled_reqs:
            not_scheduled_list.append({
                'id': req.id,
                'company_name': req.company_name,
                'company_code': req.company_code,
                'requirement_date': req.requirement_date,
                'escalation': req.escalation or '',
                'schedule_status': req.schedule_status
            })

        # Prepare pending requirements list
        pending_requirements = []
        for req in pending_reqs:
            scheduled = getattr(req, 'scheduled_details', None)
            pending_requirements.append({
                'id': req.id,
                'company_name': req.company_name,
                'company_code': req.company_code,
                'schedule_date': req.schedule_date,
                'scheduled_details': scheduled,
            })

        details.update({
            'total_requirements': total_reqs,
            'scheduled_companies': scheduled_companies,
            'not_scheduled_companies': not_scheduled_companies,
            'pending_companies': pending_companies,
            'escalated_companies': escalated_companies,
            'all_rounds_completed': all_rounds_completed,
            'placements': placements,
            'company_status': [],
            'not_scheduled_list': not_scheduled_list,
            'pending_requirements': pending_requirements,
        })

        # Company-wise status
        for req in scheduled_reqs:
            scheduled = getattr(req, 'scheduled_details', None)
            details['company_status'].append({
                'id': req.id,
                'company': req.company_name,
                'company_code': req.company_code,
                'status': req.get_schedule_status_display(),
                'scheduled_date': req.schedule_date,
                'result': scheduled.get_result_display() if scheduled else 'No Update',
                'students_appeared': scheduled.students_appeared.count() if (scheduled and hasattr(scheduled, 'students_appeared')) else 0,
                'students_selected': RequirementStudent.objects.filter(requirement=req, status='selected').count(),
                'escalation': req.escalation or ''
            })

        return details

    def get_cross_month_requirements(self, current_month):
        from datetime import datetime, timedelta
        """Requirements posted last month but scheduled this month"""
        last_month = (current_month.replace(day=1) - timedelta(days=1)).replace(day=1)
        qs = Requirement.objects.filter(
            requirement_date__year=last_month.year,
            requirement_date__month=last_month.month,
            schedule_date__year=current_month.year,
            schedule_date__month=current_month.month
        ).select_related('scheduled_details').annotate(
            students_selected_count=Count(
                'requirementstudent',
                filter=Q(requirementstudent__status='selected'),
                distinct=True
            )
        )
        return qs

from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from .models import Requirement, ScheduledRequirement
from django.utils import timezone

def mark_as_scheduled(request, requirement_id):
    requirement = get_object_or_404(Requirement, id=requirement_id)

    if request.method == 'POST':
        result_status = request.POST.get('result_status')  # Get the result status from the form
        
        # Update the requirement
        requirement.is_scheduled = True
        requirement.schedule_status = 'scheduled'
        requirement.schedule_date = timezone.now().date()
        requirement.result_status = result_status  # Save the selected result status
        requirement.save()

        # Create or update ScheduledRequirement
        ScheduledRequirement.objects.get_or_create(
            requirement=requirement,
            defaults={'scheduled_date': requirement.schedule_date}
        )
        
        return redirect('student_data:monthly_reports')

    # If GET request, show a form to pick result status
    return render(request, 'mark_as_scheduled.html', {'requirement': requirement})



# Excel export view
from django.views import View

class ExportHRReportExcelView(View):
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HR Report"
        ws.append(["Company", "Code", "Status", "Scheduled Date", "Result", "Students Appeared", "Students Selected", "Escalation"])

        requirements = Requirement.objects.exclude(schedule_date__isnull=True).select_related('scheduled_details')
        for req in requirements:
            scheduled = getattr(req, 'scheduled_details', None)
            result = scheduled.get_result_display() if scheduled else 'No Update'
            students_appeared = scheduled.students_appeared.count() if scheduled else 0
            students_selected = RequirementStudent.objects.filter(requirement=req, status='selected').count()
            escalation = req.escalation or ''
            ws.append([
                req.company_name,
                req.company_code,
                req.get_schedule_status_display(),
                req.schedule_date,
                result,
                students_appeared,
                students_selected,
                escalation
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename = f"HR_Report_{now().strftime('%Y_%m')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

from django.shortcuts import render, redirect
from django.contrib import messages
  # Import your actual model

def update_escalation(request, pk):
    if request.method == 'POST':
        # Get the requirement object
        requirement = get_object_or_404(Requirement, pk=pk)
        
        # Get form data
        escalation_reason = request.POST.get('reason')
        escalation_priority = request.POST.get('priority')
        
        # Update the requirement with escalation info
        requirement.escalation = f"{escalation_priority}: {escalation_reason}"
        requirement.save()
        
        messages.success(request, f"Escalation raised for {requirement.company_name}")
        return redirect('student_data:monthly_reports')  # Adjust to your actual redirect URL
    
    # If not POST, redirect back
    return redirect('student_data:monthly_reports')

from django.views.generic.edit import UpdateView
from django.urls import reverse_lazy
from .models import Student

from django.views.generic.edit import UpdateView
from django.urls import reverse_lazy
from .models import Student

class StudentUpdateView(UpdateView):
    model = Student
    template_name = 'student_edit.html'
    success_url = reverse_lazy('student_data:student_list')
    
    # Include all fields from the model
    fields = [
        'name', 'contact_number', 'degree', 'stream', 'yop',
        'tenth_percent', 'twelfth_percent', 'degree_percent',
        'gender', 'type_of_data', 'is_placed', 'is_dropout',
        'dropout_date', 'dropout_reason', 'overall_technical_rating',
        'total_requirements', 'scheduled_requirements'
    ]
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        
        # Customize form fields as needed
        form.fields['name'].widget.attrs.update({'class': 'form-control'})
        form.fields['contact_number'].widget.attrs.update({'class': 'form-control'})
        form.fields['degree'].widget.attrs.update({'class': 'form-control'})
        form.fields['stream'].widget.attrs.update({'class': 'form-control'})
        form.fields['yop'].widget.attrs.update({'class': 'form-control'})
        form.fields['tenth_percent'].widget.attrs.update({
            'class': 'form-control',
            'step': '0.01',
            'min': '0',
            'max': '100'
        })
        form.fields['twelfth_percent'].widget.attrs.update({
            'class': 'form-control',
            'step': '0.01',
            'min': '0',
            'max': '100'
        })
        form.fields['degree_percent'].widget.attrs.update({
            'class': 'form-control',
            'step': '0.01',
            'min': '0',
            'max': '100'
        })
        form.fields['gender'].widget.attrs.update({'class': 'form-control'})
        form.fields['type_of_data'].widget.attrs.update({'class': 'form-control'})
        form.fields['overall_technical_rating'].widget.attrs.update({'class': 'form-control'})
        form.fields['dropout_date'].widget.attrs.update({
            'class': 'form-control',
            'type': 'date'
        })
        form.fields['dropout_reason'].widget.attrs.update({
            'class': 'form-control',
            'rows': 3
        })
        
        # Conditional logic for dropout fields
        if not form.instance.is_dropout:
            form.fields['dropout_date'].disabled = True
            form.fields['dropout_reason'].disabled = True
        
        # Read-only fields that shouldn't be manually edited
        form.fields['total_requirements'].disabled = True
        form.fields['scheduled_requirements'].disabled = True
       # form.fields['created_at'].disabled = True  # If you choose to include it
        
        return form
    
    def form_valid(self, form):
        # Custom save logic if needed
        response = super().form_valid(form)
        
        # Update derived fields
        self.object.update_overall_technical_rating()
        self.object.update_requirement_counts()
        self.object.update_placement_status()
        
        return response  # or wherever you want to redirect after edit



from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.serializers import serialize
import csv
from .models import GotPlacedOutside, Student
from .forms import GotPlacedOutsideForm

def add_got_placed_student(request):
    search_query = request.GET.get('search_query', '')
    students = Student.objects.filter(is_placed=False)

    if search_query:
        # Search by name OR mobile number
        students = students.filter(name__icontains=search_query) | students.filter(contact_number__icontains=search_query)

    form = GotPlacedOutsideForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        placement = form.save()
        student = placement.student
        student.is_placed = True
        student.save()
        messages.success(request, 'Student placement details have been added successfully.')
        return redirect('student_data:placed_students')

    context = {
        'form': form,
        'students': students,
        'search_query': search_query
    }

    return render(request, 'got_placed_outside.html', context)


# AJAX handler for live search
def ajax_search_students(request):
    query = request.GET.get('q', '')
    students = Student.objects.filter(is_placed=False)
    if query:
        students = students.filter(name__icontains=query) | students.filter(contact_number__icontains=query)

    data = list(students.values('id', 'name', 'contact_number'))
    return JsonResponse({'students': data})


# CSV Export of current search results
def export_students_csv(request):
    search_query = request.GET.get('search_query', '')
    students = Student.objects.filter(is_placed=False)

    if search_query:
        students = students.filter(name__icontains=search_query) | students.filter(contact_number__icontains=search_query)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="students.csv"'

    writer = csv.writer(response)
    writer.writerow(['Name', 'Mobile Number'])

    for student in students:
        writer.writerow([student.name, student.contact_number])

    return response
import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from django.views.generic import FormView
from django.urls import reverse_lazy
from django.core.exceptions import ObjectDoesNotExist
from .models import Student, GotPlacedOutside
from .forms import BulkOutsidePlacementForm
from django.utils import timezone
import re


class BulkOutsidePlacementView(FormView):
    template_name = 'bulk_outside_placement.html'
    form_class = BulkOutsidePlacementForm
    success_url = reverse_lazy('student_data:bulk_outside_placement')  # Use reverse_lazy if needed in class attr

    def get(self, request, *args, **kwargs):
        if 'bulk_upload_errors' in self.request.session:
            del self.request.session['bulk_upload_errors']
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        excel_file = form.cleaned_data['excel_file']

        try:
            df = pd.read_excel(excel_file)
            df.columns = df.columns.str.lower()  # Normalize column names

            success_count = 0
            error_count = 0
            errors = []

            for index, row in df.iterrows():
                try:
                    mobile_number = str(row.get('mobile_number', '')).strip()

                    if not mobile_number:
                        raise ValueError("Mobile number is required")

                    if not re.fullmatch(r'\d{10}', mobile_number):
                        raise ValueError(f"Invalid mobile number format: {mobile_number}. Must be a 10-digit number.")

                    try:
                        student = Student.objects.get(contact_number=mobile_number)
                    except Student.DoesNotExist:
                        raise ValueError(f"No student found with mobile number: {mobile_number}")

                    if student.is_placed and not GotPlacedOutside.objects.filter(student=student).exists():
                        raise ValueError(f"Student {student.name} is already placed internally and cannot be added externally.")

                    placement, created = GotPlacedOutside.objects.get_or_create(
                        student=student,
                        defaults={
                            'company_name': row.get('company_name', ''),
                            'package': row.get('package', ''),
                            'role': row.get('role', ''),
                            'placed_date': row.get('placed_date', timezone.now())
                        }
                    )

                    if not created:
                        placement.company_name = row.get('company_name', placement.company_name)
                        placement.package = row.get('package', placement.package)
                        placement.role = row.get('role', placement.role)
                        placement.placed_date = row.get('placed_date', placement.placed_date)
                        placement.save()

                    if created:
                        student.is_placed = True
                        student.save()

                    success_count += 1

                except ValueError as ve:
                    error_count += 1
                    errors.append(f"Row {index + 2}: {str(ve)}")
                except Exception as e:
                    error_count += 1
                    errors.append(f"Row {index + 2}: Unexpected error - {str(e)}")

            message = f"Successfully processed {success_count} records."
            if error_count > 0:
                message += f" {error_count} records had errors."
                messages.warning(self.request, message)
                self.request.session['bulk_upload_errors'] = errors
            else:
                messages.success(self.request, message)

            return super().form_valid(form)

        except Exception as e:
            messages.error(self.request, f"Error processing file: {str(e)}")
            return self.form_invalid(form)
# views.py
from django.http import HttpResponse
import pandas as pd
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO

def download_sample_excel(request):
    # Create a workbook and add worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample"
    
    # Add headers
    headers = ['mobile_number', 'company_name', 'package', 'role', 'placed_date']
    ws.append(headers)
    
    # Add sample data
    sample_data = [
        ['9876543210', 'Company A', '10 LPA', 'Software Engineer', '2023-01-15'],
        ['8765432109', 'Company B', '12 LPA', 'Data Scientist', '2023-02-20']
    ]
    for row in sample_data:
        ws.append(row)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=sample_outside_placement.xlsx'
    return response


from django.shortcuts import render
from .models import GotPlacedOutside, Student

from django.shortcuts import render
from django.db.models import Q
from django.core.paginator import Paginator
from .models import GotPlacedOutside

def students_placed_outside(request):
    # Get search query
    query = request.GET.get('q', '')
    
    # Filter students based on search query if provided
    if query:
        students_placed = GotPlacedOutside.objects.filter(
            Q(student__name__icontains=query) |
            Q(company_name__icontains=query) |
            Q(role__icontains=query) |
            Q(student__degree__icontains=query) |
            Q(student__type_of_data__icontains=query)
        )
    else:
        # Get all students who have been placed outside
        students_placed = GotPlacedOutside.objects.all()
    
    # Order by most recent placements first
    students_placed = students_placed.order_by('-placed_date')
    
    # Set up pagination - 12 students per page
    paginator = Paginator(students_placed, 12)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Create context for rendering the template
    context = {
        'page_obj': page_obj,
    }

    return render(request, 'students_placed_outside.html', context)



from django.shortcuts import get_object_or_404, redirect, render
from django.views.generic import View
from django.contrib import messages
from django.urls import reverse
from .models import Student, Subject, StudentSubjectRating
from .forms import StudentSubjectRatingFormSet

class StudentSubjectRatingsUpdateView(View):
    template_name = 'students/update_subject_ratings.html'
    
    def get(self, request, student_id):
        student = get_object_or_404(Student, pk=student_id)
        existing_ratings = student.subject_ratings.select_related('subject')
        
        # Initialize formset with existing data
        formset = StudentSubjectRatingFormSet(
            queryset=existing_ratings,
            prefix='ratings'
        )
        
        # Add forms for subjects that don't have ratings yet
        existing_subject_ids = existing_ratings.values_list('subject_id', flat=True)
        missing_subjects = Subject.objects.filter(is_active=True).exclude(id__in=existing_subject_ids)
        
        for subject in missing_subjects:
            formset.extra += 1
            formset.forms.append(formset.empty_form)
            formset.forms[-1].initial = {
                'subject': subject,
                'student': student
            }
        
        context = {
            'student': student,
            'formset': formset,
        }
        return render(request, self.template_name, context)
    
    def post(self, request, student_id):
        student = get_object_or_404(Student, pk=student_id)
        formset = StudentSubjectRatingFormSet(
            request.POST,
            prefix='ratings'
        )
        
        if formset.is_valid():
            instances = formset.save(commit=False)
            
            for instance in instances:
                instance.student = student
                instance.save()
            
            # Delete any ratings that were marked for deletion
            for obj in formset.deleted_objects:
                obj.delete()
            
            # Update student's overall technical rating
            student.save()
            
            messages.success(request, 'Subject ratings updated successfully!')
            return redirect(reverse('student-detail', kwargs={'pk': student_id}))
        
        context = {
            'student': student,
            'formset': formset,
        }
        return render(request, self.template_name, context)
    
#import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from django.views.generic import View
from .models import Student, Subject, StudentSubjectRating
from .forms import BulkStudentRatingUploadForm
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

class BulkStudentRatingUploadView(View):
    template_name = 'bulk_rating_upload.html'
    form_class = BulkStudentRatingUploadForm
    
    def get(self, request):
        subjects = Subject.objects.filter(is_active=True)
        rating_choices = StudentSubjectRating.RATING_CHOICES
        return render(request, self.template_name, {
            'form': self.form_class(),
            'subjects': subjects,
            'rating_choices': rating_choices
        })
    
    def post(self, request):
        form = self.form_class(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form})
    
        try:
            file = request.FILES['file']
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
        
        # Convert all data to strings and strip whitespace
            df = df.applymap(lambda x: str(x).strip() if pd.notna(x) else '')
        
        # Process each row
            success_count = 0
            error_count = 0
            errors = []
            valid_subjects = set(Subject.objects.filter(is_active=True).values_list('name', flat=True))
            valid_ratings = set(dict(StudentSubjectRating.RATING_CHOICES).keys())
        
            for index, row in df.iterrows():
                row_num = index + 2  # Excel rows start at 1, header is row 1
                try:
                    mobile_no = row[0]
                    if not mobile_no:
                        errors.append(f"Row {row_num}: Mobile number is required")
                        error_count += 1
                        continue
                
                # Get student
                    student = Student.objects.filter(contact_number=mobile_no).first()
                    if not student:
                        errors.append(f"Row {row_num}: Student with mobile {mobile_no} not found")
                        error_count += 1
                        continue
                
                # Process subject-rating pairs
                    subjects_processed = 0
                    subject_rating_pairs = []
                
                    for i in range(1, min(len(row), 10), 2):  # Max 5 subjects (columns 1-9)
                        if i+1 >= len(row):
                            break
                        
                        subject_name = row[i]
                        rating = row[i+1].lower() if pd.notna(row[i+1]) else ''
                    
                    # Skip empty pairs
                        if not subject_name and not rating:
                            continue
                        
                    # Validate subject
                        if not subject_name:
                            errors.append(f"Row {row_num}: Subject name is missing for pair {i//2 + 1}")
                            continue
                        
                        if subject_name not in valid_subjects:
                            errors.append(f"Row {row_num}: Subject '{subject_name}' not found. Valid subjects are: {', '.join(valid_subjects)}")
                            continue
                        
                    # Validate rating
                        if not rating:
                            errors.append(f"Row {row_num}: Rating is missing for subject '{subject_name}'")
                            continue
                        
                        if rating not in valid_ratings:
                            errors.append(f"Row {row_num}: Invalid rating '{rating}' for subject '{subject_name}'. Valid ratings are: {', '.join(valid_ratings)}")
                            continue
                        
                        subject_rating_pairs.append((subject_name, rating))
                        subjects_processed += 1
                
                # Validate minimum subjects
                    if subjects_processed < 2:
                        errors.append(f"Row {row_num}: Minimum 2 subjects required, found {subjects_processed}")
                        error_count += 1
                        continue
                
                # Save all valid pairs for this student
                    for subject_name, rating in subject_rating_pairs:
                        subject = Subject.objects.get(name=subject_name)
                        StudentSubjectRating.objects.update_or_create(
                            student=student,
                        subject=subject,
                        defaults={'rating': rating}
                    )
                
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: Error processing - {str(e)}")
                    error_count += 1
        
        # Prepare result message
            msg = f"Successfully processed {success_count} students"
            if error_count > 0:
                msg += f", with {error_count} errors"
        
            messages.success(request, msg)
            if errors:
                messages.error(request, "Errors encountered during processing:")
                for error in errors[:10]:  # Show first 10 errors to avoid flooding
                    messages.error(request, error)
                if len(errors) > 10:
                    messages.warning(request, f"... and {len(errors)-10} more errors not shown")
        
            return redirect('student_data:bulk_rating_upload')
        
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
            return render(request, self.template_name, {'form': form})
class DownloadRatingTemplateView(View):
    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Ratings Template"
        
        # Header row
        headers = ["MobileNo (Required)", "Subject1 (Required)", "Subject1_Rating (Required)", 
                  "Subject2 (Required)", "Subject2_Rating (Required)",
                  "Subject3 (Optional)", "Subject3_Rating (Optional)",
                  "Subject4 (Optional)", "Subject4_Rating (Optional)",
                  "Subject5 (Optional)", "Subject5_Rating (Optional)"]
        ws.append(headers)
        
        # Style and instructions
        bold_font = Font(bold=True)
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_num).font = bold_font
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20
        
        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        ws_instructions.append(["INSTRUCTIONS:"])
        ws_instructions.append(["1. Fill in student mobile numbers in first column"])
        ws_instructions.append(["2. For each student, provide at least 2 subject-rating pairs"])
        ws_instructions.append(["3. Use the dropdowns to select valid subjects and ratings"])
        ws_instructions.append(["", ""])
        ws_instructions.append(["VALID SUBJECTS:", ", ".join([s.name for s in Subject.objects.filter(is_active=True)])])
        ws_instructions.append(["VALID RATINGS:", ", ".join([r[0] for r in StudentSubjectRating.RATING_CHOICES])])
        
        # Add sample data
        sample_data = [
            "9876543210",        # MobileNo
            "core_java",        # Subject1
            "good",             # Rating1
            "sql",              # Subject2
            "excellent",        # Rating2
            "communication",    # Subject3 (optional)
            "average",         # Rating3 (optional)
            "",                 # Subject4 (optional)
            "",                 # Rating4 (optional)
            "",                 # Subject5 (optional)
            ""                  # Rating5 (optional)
        ]
        ws.append(sample_data)
        # Get active subjects and rating choices
        subjects = Subject.objects.filter(is_active=True)
        subject_names = [subj.name for subj in subjects]
        subject_choices = ",".join(subject_names)
        rating_choices = ",".join([choice[0] for choice in StudentSubjectRating.RATING_CHOICES])
        
        # Apply data validation
        for i in range(2, 12, 2):  # Subject columns (2,4,6,8,10)
            # Subject validation
            dv = DataValidation(
                type="list", 
                formula1=f'"{subject_choices}"', 
                allow_blank=(i > 6)  # Only required for first 3 subjects
            )
            ws.add_data_validation(dv)
            dv.add(f"{get_column_letter(i)}2:{get_column_letter(i)}1048576")
            
            # Rating validation (next column)
            if i < 10:  # Only up to Subject4_Rating
                dv = DataValidation(
                    type="list", 
                    formula1=f'"{rating_choices}"', 
                    allow_blank=(i > 6)  # Only required for first 3 ratings
                )
                ws.add_data_validation(dv)
                dv.add(f"{get_column_letter(i+1)}2:{get_column_letter(i+1)}1048576")
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=student_ratings_template.xlsx'
        wb.save(response)
        return response

from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from .models import Student, StudentSubjectRating
from .forms import StudentSubjectRatingForm

def add_subject_rating(request, pk):
    student = get_object_or_404(Student, pk=pk)
    rating_instance = None

    if request.method == 'POST':
        form = StudentSubjectRatingForm(request.POST)

        if form.is_valid():
            subject = form.cleaned_data['subject']

            # Try to get an existing rating for this subject & student
            rating_instance = StudentSubjectRating.objects.filter(
                student=student,
                subject=subject
            ).first()

            if rating_instance:
                form = StudentSubjectRatingForm(request.POST, instance=rating_instance)

            if form.is_valid():
                rating = form.save(commit=False)
                rating.student = student

                # Only set evaluated_by for new ratings
                if not rating_instance:
                    rating.evaluated_by = request.user.get_full_name()

                rating.save()

                action = 'updated' if rating_instance else 'added'
                messages.success(request, f'Subject rating {action} successfully!')
                return redirect('student_data:student_detail', student_id=student.id)
    else:
        form = StudentSubjectRatingForm()

    context = {
        'form': form,
        'student': student,
        'is_edit': False  # Not needed unless you're dynamically changing UI
    }
    return render(request, 'add_subject_rating.html', context)



from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
import pandas as pd
from io import BytesIO
from datetime import datetime
from .models import Student

def bulk_add_dropouts(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('dropout_file')
        
        if not excel_file:
            messages.error(request, "Please upload an Excel file")
            return redirect('bulk_add_dropouts')
        
        try:
            # Read the Excel file
            df = pd.read_excel(BytesIO(excel_file.read()))
            
            # Validate columns
            required_columns = {'mobile', 'reason'}
            if not required_columns.issubset(df.columns):
                missing = required_columns - set(df.columns)
                messages.error(request, f"Missing required columns: {', '.join(missing)}")
                return redirect('bulk_add_dropouts')
            
            # Process data
            success_count = 0
            error_messages = []
            
            with transaction.atomic():
                for index, row in df.iterrows():
                    mobile = str(row['mobile']).strip()
                    reason = str(row['reason']).strip() if pd.notna(row['reason']) else ""
                    
                    try:
                        student = Student.objects.get(contact_number=mobile)
                        
                        if student.is_dropout:
                            error_messages.append(f"Row {index+2}: {mobile} - Already marked as dropout")
                            continue
                            
                        student.is_dropout = True
                        student.dropout_date = datetime.now().date()
                        student.dropout_reason = reason
                        student.save()
                        success_count += 1
                        
                    except Student.DoesNotExist:
                        error_messages.append(f"Row {index+2}: {mobile} - Student not found")
                    except Exception as e:
                        error_messages.append(f"Row {index+2}: {mobile} - Error: {str(e)}")
            
            if success_count > 0:
                messages.success(request, f"Successfully marked {success_count} students as dropouts")
            
            if error_messages:
                messages.warning(request, f"Completed with {len(error_messages)} errors")
                request.session['bulk_upload_errors'] = error_messages
            
            return redirect('student_data:bulk_add_dropouts')
            
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
            return redirect('bulk_add_dropouts')
    
    # Get error messages from session if any
    error_messages = request.session.pop('bulk_upload_errors', [])
    
    context = {
        'error_messages': error_messages,
    }
    return render(request, 'bulk_add_dropouts.html', context)

from django.shortcuts import render
from django.utils import timezone
from datetime import datetime, time, timedelta
from .models import Requirement

def tomorrows_requirements(request):
    tomorrow = timezone.localdate() + timedelta(days=1)

    start = timezone.make_aware(datetime.combine(tomorrow, time.min))
    end = timezone.make_aware(datetime.combine(tomorrow, time.max))

    requirements = Requirement.objects.filter(
        schedule_date__range=(start, end),
        is_scheduled=True
    ).order_by('schedule_time')

    context = {
        'tomorrows_requirements': requirements,   # ✅ this name matches your template
        'tomorrow_date': tomorrow                 # ✅ used in your template header
    }

    return render(request, 'tommarows_requirement.html', context)




def students_attending_tomorrow(request):
    tomorrow = timezone.localdate() + timezone.timedelta(days=1)

    requirements_tomorrow = Requirement.objects.filter(
        schedule_date=tomorrow,
        schedule_status='scheduled'
    ).order_by('schedule_time').prefetch_related('requirementstudent_set__student')

    tomorrows_students = []

    for requirement in requirements_tomorrow:
        for rs in requirement.requirementstudent_set.all():
            student = rs.student
            tomorrows_students.append({
                'id': student.id,
                'name': student.name,
                'contact_number': student.contact_number,
                'degree': student.degree,
                'stream': student.stream,
                'yop': student.yop,
                'degree_percent': student.degree_percent,
                'twelfth_percent': student.twelfth_percent,
                'tenth_percent': student.tenth_percent,
                'overall_technical_rating': student.overall_technical_rating,
                'is_dropout': student.is_dropout,
                'dropout_reason': student.dropout_reason,
                'is_placed': student.is_placed,
                # Add company information to each student
                'requirement': {
                    'company_name': requirement.company_name,
                }
            })

    context = {
        'tomorrows_students': tomorrows_students,
        'tomorrow_date': tomorrow,
    }

    return render(request, 'students_attending_tomorrow.html', context)





import io
import zipfile
import pandas as pd
from django.http import HttpResponse
from django.utils import timezone
from .models import RequirementStudent, Requirement, Student
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

def export_tomorrow_scheduled_requirements(request):
    tomorrow = timezone.now().date() + timezone.timedelta(days=1)
    requirements = Requirement.objects.filter(schedule_date=tomorrow, schedule_status='scheduled')
    
    if not requirements.exists():
        return HttpResponse("No scheduled requirements for tomorrow.", status=404)

    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for req in requirements:
            req_students = RequirementStudent.objects.filter(
                requirement=req
            ).select_related('student').order_by('student__name')
            
            student_data = []
            for req_student in req_students:
                student = req_student.student
                student_data.append({
                    'Student Name': student.name,
                    'Contact Number': student.contact_number,
                    'Gender': student.gender,
                    'Degree': student.degree,
                    'Stream': student.stream,
                    'Year of Passing': student.yop,
                    '10th %': student.tenth_percent,
                    '12th %': student.twelfth_percent,
                    'Degree %': student.degree_percent,
                    'Type of Data': student.type_of_data,
                    'Status': req_student.status,
                    'Feedback': req_student.feedback or '',
                    'Is Placed': "Yes" if student.is_placed else "No"
                })

            student_df = pd.DataFrame(student_data)

            requirement_details = {
                'Requirement Field': [
                    'Company Name',
                    'Company Code',
                    'Schedule Date',
                    'Schedule Time',
                    'Requirement Date',
                    'Description',
                    'Schedule Status',
                    'Escalation'
                ],
                'Value': [
                    req.company_name,
                    req.company_code,
                    str(req.schedule_date),
                    str(req.schedule_time),
                    str(req.requirement_date),
                    req.description,
                    req.schedule_status,
                    req.escalation
                ]
            }
            details_df = pd.DataFrame(requirement_details)

            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                details_df.to_excel(writer, index=False, sheet_name='Requirement Details')
                student_df.to_excel(writer, index=False, sheet_name='Students')

                workbook = writer.book
                details_sheet = writer.sheets['Requirement Details']
                students_sheet = writer.sheets['Students']

                # Format both sheets
                header_font = Font(bold=True)
                header_fill = PatternFill("solid", fgColor="DDDDDD")

                # Format Requirement Details Sheet
                for col in range(1, 3):  # A and B columns
                    col_letter = get_column_letter(col)
                    details_sheet.column_dimensions[col_letter].width = 30
                    cell = details_sheet[f"{col_letter}1"]
                    cell.font = header_font
                    cell.fill = header_fill

                # Format Students Sheet
                for col_idx, column in enumerate(student_df.columns, 1):
                    max_length = max(
                        student_df[column].astype(str).map(len).max(),
                        len(column)
                    )
                    col_letter = get_column_letter(col_idx)
                    students_sheet.column_dimensions[col_letter].width = min(max_length + 2, 30)
                    header_cell = students_sheet[f"{col_letter}1"]
                    header_cell.font = header_font
                    header_cell.fill = header_fill

            excel_buffer.seek(0)
            filename = f"{req.company_name.replace(' ', '_')}_{req.id}_{tomorrow}.xlsx"
            zip_file.writestr(filename, excel_buffer.read())

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="scheduled_requirements_{tomorrow}.zip"'
    return response




from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from .models import Student, RequirementStudent, GotPlacedOutside
from datetime import datetime

def update_placement_status(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    
    if request.method == 'POST':
        placement_type = request.POST.get('placement_type')
        
        if placement_type == 'internal':
            req_student_id = request.POST.get('selected_requirement')
            if req_student_id:
                # Get selected requirement with optimized query
                selected_req = get_object_or_404(
                    RequirementStudent.objects.select_related('requirement'),
                    id=req_student_id, 
                    student=student
                )
                selected_req.status = 'selected'
                selected_req.save()
                
                # Reject all other requirements for this student
                RequirementStudent.objects.filter(student=student).exclude(id=req_student_id).update(status='rejected')
                
                # Critical addition: Reject other students' pending applications for the same requirement
                RequirementStudent.objects.filter(
                    requirement=selected_req.requirement,
                    status='pending'
                ).exclude(student=student).update(status='rejected')
            
            # Clear external placement if exists
            GotPlacedOutside.objects.filter(student=student).delete()
            
        elif placement_type == 'external':
            company_name = request.POST.get('company_name')
            package = request.POST.get('package')
            role = request.POST.get('role')
            placed_date = request.POST.get('placed_date') or timezone.now().date()
            
            # Validate date format
            try:
                placed_date = datetime.strptime(placed_date, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                placed_date = timezone.now().date()
            
            # Update or create external placement
            GotPlacedOutside.objects.update_or_create(
                student=student,
                defaults={
                    'company_name': company_name,
                    'package': package,
                    'role': role,
                    'placed_date': placed_date
                }
            )
            
            # Reject all internal requirements
            RequirementStudent.objects.filter(student=student).update(status='rejected')
        
        # Update student status and redirect
        student.update_placement_status()
        return redirect('student_data:student_detail', student_id=student.id)
    
    else:
        # GET request with optimized queries
        internal_reqs = student.requirementstudent_set.select_related('requirement').all()
        external_placement = getattr(student, 'placed_outside', None)
        
        # Determine initial selection state
        initial_type = 'external' if external_placement else \
                      'internal' if internal_reqs.filter(status='selected').exists() else None
        
        context = {
            'student': student,
            'internal_reqs': internal_reqs,
            'external_placement': external_placement,
            'initial_type': initial_type,
        }
        return render(request, 'update_placement.html', context)
    

from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from .models import Student, RequirementStudent, GotPlacedOutside

def remove_placement(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    
    # Try to find existing internal placement
    internal_placement = RequirementStudent.objects.filter(
        student=student,
        status='selected'
    ).first()

    if internal_placement:
        # Reject the selected internal placement
        internal_placement.status = 'rejected'
        internal_placement.save()
        
        # Delete any external placement record if exists
        GotPlacedOutside.objects.filter(student=student).delete()
        
        messages.success(request, 
            f"Internal placement with {internal_placement.requirement.company_name} "
            f"has been removed and status set to rejected.")
            
    else:
        # Try to remove external placement
        external_placement = GotPlacedOutside.objects.filter(student=student).first()
        if external_placement:
            company_name = external_placement.company_name
            external_placement.delete()
            messages.success(request, 
                f"External placement with {company_name} has been removed.")
        else:
            messages.warning(request, "No active placement found to remove.")
    
    # Update student's placement status
    student.update_placement_status()
    
    return redirect('student_data:student_detail', student_id=student.id)

from django.http import JsonResponse
from django.views import View
from django.utils import timezone
from .models import Requirement, RequirementStudent, Student
from django.urls import reverse


# ---------------------------
# AJAX Views
# ---------------------------

class TodaysRequirementsAjaxView(View):
    def get(self, request, *args, **kwargs):
        today = timezone.now().date()
        requirements = Requirement.objects.filter(schedule_date=today)
        data = [
            {
                "company_name": req.company_name,
                "company_code": req.company_code or "",
                "schedule_time": req.schedule_time.strftime("%I:%M %p") if req.schedule_time else "--",
                "detail_url": request.build_absolute_uri(req.get_absolute_url()),
            }
            for req in requirements
        ]
        return JsonResponse({"requirements": data})


class TomorrowsRequirementsAjaxView(View):
    def get(self, request, *args, **kwargs):
        tomorrow = timezone.now().date() + timezone.timedelta(days=1)
        requirements = Requirement.objects.filter(schedule_date=tomorrow)
        data = [
            {
                "company_name": req.company_name,
                "company_code": req.company_code or "",
                "schedule_time": req.schedule_time.strftime("%I:%M %p") if req.schedule_time else "--",
                "detail_url": request.build_absolute_uri(req.get_absolute_url()),
            }
            for req in requirements
        ]
        return JsonResponse({"requirements": data})


class TodaysStudentsAjaxView(View):
    def get(self, request, *args, **kwargs):
        today = timezone.now().date()
        student_requirements = RequirementStudent.objects.filter(
            requirement__schedule_date=today
        ).select_related('student', 'requirement')

        data = [
            {
                "student_name": sr.student.name,
                "degree": sr.student.degree,
                "stream": sr.student.stream,
                "company_name": sr.requirement.company_name,
                "student_detail_url": request.build_absolute_uri(sr.student.get_absolute_url()),
                "requirement_detail_url": request.build_absolute_uri(sr.requirement.get_absolute_url()),
            }
            for sr in student_requirements
        ]

        return JsonResponse({"students": data})


class TomorrowsStudentsAjaxView(View):
    def get(self, request, *args, **kwargs):
        tomorrow = timezone.now().date() + timezone.timedelta(days=1)
        student_requirements = RequirementStudent.objects.filter(
            requirement__schedule_date=tomorrow
        ).select_related('student', 'requirement')

        data = [
            {
                "student_name": sr.student.name,
                "degree": sr.student.degree,
                "stream": sr.student.stream,
                "company_name": sr.requirement.company_name,
                "student_detail_url": request.build_absolute_uri(sr.student.get_absolute_url()),
                "requirement_detail_url": request.build_absolute_uri(sr.requirement.get_absolute_url()),
            }
            for sr in student_requirements
        ]

        return JsonResponse({"students": data})